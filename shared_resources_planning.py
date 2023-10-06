import os
import time
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
from math import sqrt, isclose
import pyomo.opt as po
import pyomo.environ as pe
from network_planning import NetworkPlanning
from shared_energy_storage import SharedEnergyStorage
from shared_energy_storage_data import SharedEnergyStorageData
from planning_parameters import PlanningParameters
from helper_functions import *


# ======================================================================================================================
#   Class SHARED RESOURCES PLANNING
# ======================================================================================================================
class SharedResourcesPlanning:

    def __init__(self, data_dir, filename):
        self.name = filename.replace('.txt', '')
        self.data_dir = data_dir
        self.filename = filename
        self.market_data_file = str()
        self.results_dir = os.path.join(data_dir, 'Results')
        self.plots_dir = os.path.join(data_dir, 'Results', 'Plots')
        self.diagrams_dir = os.path.join(data_dir, 'Diagrams')
        self.params_file = str()
        self.years = dict()
        self.days = dict()
        self.num_instants = 0
        self.discount_factor = 0.00
        self.cost_energy_p = dict()
        self.cost_secondary_reserve = dict()
        self.cost_tertiary_reserve_up = dict()
        self.cost_tertiary_reserve_down = dict()
        self.prob_market_scenarios = list()
        self.distribution_networks = dict()
        self.transmission_network = NetworkPlanning()
        self.shared_ess_data = SharedEnergyStorageData()
        self.params = PlanningParameters()

    def run_planning_problem(self):
        print('[INFO] Running PLANNING PROBLEM...')
        _run_planning_problem(self)

    def run_without_coordination(self):
        print('[INFO] Running PLANNING PROBLEM WITHOUT COORDINATION...')
        _run_operational_planning_without_coordination(self)

    def run_operational_planning(self, candidate_solution=dict()):
        print('[INFO] Running OPERATIONAL PLANNING...')
        if not candidate_solution:
            candidate_solution = self.get_initial_candidate_solution()
        return _run_operational_planning(self, candidate_solution)

    def update_models_with_candidate_solution(self, tso_model, dso_models, esso_model, candidate_solution):
        self.transmission_network.update_model_with_candidate_solution(tso_model, candidate_solution['total_capacity'])
        for node_id in self.active_distribution_network_nodes:
            self.distribution_networks[node_id].update_model_with_candidate_solution(dso_models[node_id], candidate_solution['total_capacity'])
        self.shared_ess_data.update_model_with_candidate_solution(esso_model, candidate_solution['investment'])

    def update_admm_consensus_variables(self, tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, params):
        _update_admm_consensus_variables(self, tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, params)

    def read_planning_problem(self):
        _read_planning_problem(self)

    def read_market_data_from_file(self):
        _read_market_data_from_file(self)

    def read_planning_parameters_from_file(self):
        print(f'[INFO] Reading PLANNING PARAMETERS from file {self.params_file} ...')
        filename = os.path.join(self.data_dir, self.params_file)
        self.params.read_parameters_from_file(filename)

    def plot_diagram(self):
        _plot_networkx_diagram(self)

    def get_initial_candidate_solution(self):
        return _get_initial_candidate_solution(self)

    def write_operational_planning_results_to_excel(self, tso_model, dso_models, esso_model, results, primal_evolution=list()):
        filename = self.filename.replace('.txt', '') + '_operational_planning_results'
        processed_results = _process_operational_planning_results(self, tso_model, dso_models, esso_model, results)
        _write_operational_planning_results_to_excel(self, processed_results, primal_evolution=primal_evolution, filename=filename)

    def write_operational_planning_results_without_coordination_to_excel(self, tso_model, dso_models, results):
        filename = self.filename.replace('.txt', '') + '_operational_planning_results_no_coordination'
        processed_results = _process_operational_planning_results_no_coordination(self, tso_model, dso_models, results)
        _write_operational_planning_results_no_coordination_to_excel(self, processed_results, filename)

    def write_planning_results_to_excel(self, operational_planning_models, operational_results=dict(), bound_evolution=dict(), execution_time='N/A'):
        tso_model = operational_planning_models['tso']
        dso_models = operational_planning_models['dso']
        esso_model = operational_planning_models['esso']
        filename = self.filename.replace('.txt', '') + '_planning_results'
        shared_ess_capacity = self.shared_ess_data.get_investment_and_available_capacities(esso_model)
        shared_ess_processed_results = self.shared_ess_data.process_results(esso_model, execution_time=execution_time)
        if operational_results['tso']:
            operational_planning_processed_results = _process_operational_planning_results(self, tso_model, dso_models, esso_model, operational_results)
            _write_planning_results_to_excel(self, shared_ess_processed_results, shared_ess_capacity, operational_planning_processed_results, bound_evolution, filename)
        else:
            _write_planning_results_to_excel(self, shared_ess_processed_results, shared_ess_capacity, bound_evolution=bound_evolution, filename=filename)


# ======================================================================================================================
#  PLANNING functions
# ======================================================================================================================
def _run_planning_problem(planning_problem):

    shared_ess_data = planning_problem.shared_ess_data
    shared_ess_parameters = shared_ess_data.params
    benders_parameters = planning_problem.params.benders

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization
    iter = 1
    convergence = False
    lower_bound = -shared_ess_parameters.budget * 1e3
    upper_bound = shared_ess_parameters.budget * 1e3
    lower_bound_evolution = [lower_bound]
    upper_bound_evolution = [upper_bound]
    candidate_solution = planning_problem.get_initial_candidate_solution()
    #candidate_solution = shared_ess_data.get_initial_candidate_solution()

    start = time.time()
    esso_master_problem_model = shared_ess_data.build_master_problem()

    # Benders' main cycle
    while iter < benders_parameters.num_max_iters and not convergence:

        print(f'=============================================== ITERATION #{iter} ==============================================')
        print(f'[INFO] Iter {iter}. LB = {lower_bound}, UB = {upper_bound}')

        _print_candidate_solution(candidate_solution)

        # 1. Subproblem
        # 1.1. Solve operational planning, with fixed investment variables,
        # 1.2. Get coupling constraints' sensitivities (subproblem)
        # 1.3. Get OF value (upper bound) from the subproblem
        operational_results, sensitivities, lower_level_models = planning_problem.run_operational_planning(candidate_solution)
        upper_bound = shared_ess_data.compute_primal_value(lower_level_models['esso'])
        upper_bound_evolution.append(upper_bound)
        print('[INFO] ESSO. Estimated profit: {:.6f}'.format(-upper_bound))

        #  - Convergence check
        if isclose(upper_bound, lower_bound, abs_tol=benders_parameters.tol_abs, rel_tol=benders_parameters.tol_rel):
            lower_bound_evolution.append(lower_bound)
            convergence = True
            break

        iter += 1

        # 2. Solve Master problem
        # 2.1. Add Benders' cut, based on the sensitivities obtained from the subproblem
        # 2.2. Run master problem optimization
        # 2.3. Get new capacity values, and the value of alpha (lower bound)
        shared_ess_data.add_benders_cut(esso_master_problem_model, upper_bound, sensitivities, candidate_solution['investment'])
        shared_ess_data.optimize(esso_master_problem_model)
        candidate_solution = shared_ess_data.get_candidate_solution(esso_master_problem_model)
        lower_bound = pe.value(esso_master_problem_model.alpha)
        lower_bound_evolution.append(lower_bound)

        #  - Convergence check
        if isclose(upper_bound, lower_bound, abs_tol=benders_parameters.tol_abs, rel_tol=benders_parameters.tol_rel):
            lower_bound_evolution.append(lower_bound)
            convergence = True
            break

    if not convergence:
        print('[WARNING] Convergence not obtained!')

    print('[INFO] Final. LB = {}, UB = {}'.format(lower_bound, upper_bound))

    # Write results
    end = time.time()
    total_execution_time = end - start
    bound_evolution = {'lower_bound': lower_bound_evolution, 'upper_bound': upper_bound_evolution}
    planning_problem.write_planning_results_to_excel(lower_level_models, operational_results, bound_evolution, execution_time=total_execution_time)


def _print_candidate_solution(candidate_solution):

    print('[INFO] Candidate solution:')

    # Header
    print('\t\t{:3}\t{:10}\t'.format('', 'Capacity'), end='')
    for node_id in candidate_solution['total_capacity']:
        for year in candidate_solution['total_capacity'][node_id]:
            print(f'{year}\t', end='')
        print()
        break

    # Values
    for node_id in candidate_solution['total_capacity']:
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'S, [MVA]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['s']), end='')
        print()
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'E, [MVAh]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['e']), end='')
        print()


# ======================================================================================================================
#  OPERATIONAL PLANNING functions
# ======================================================================================================================
def _run_operational_planning(planning_problem, candidate_solution):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data
    admm_parameters = planning_problem.params.admm
    results = {'tso': dict(), 'dso': dict(), 'esso': dict()}

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization

    print('[INFO]\t\t - Initializing...')

    #start = time.time()
    primal_evolution = list()

    # Create ADMM variables
    consensus_vars, dual_vars, consensus_vars_prev_iter = create_admm_variables(planning_problem)

    # Create Operational Planning models
    dso_models = create_distribution_networks_models(distribution_networks, consensus_vars['interface']['pf']['dso'], consensus_vars['ess']['dso'], candidate_solution['total_capacity'])
    update_distribution_models_to_admm(distribution_networks, dso_models, consensus_vars['interface']['pf']['dso'], admm_parameters)

    tso_model = create_transmission_network_model(transmission_network, consensus_vars['interface']['v'], consensus_vars['interface']['pf'], consensus_vars['ess']['tso'], candidate_solution['total_capacity'])
    update_transmission_model_to_admm(transmission_network, tso_model, consensus_vars['interface']['pf'], admm_parameters)

    esso_model = create_shared_energy_storage_model(shared_ess_data, candidate_solution['investment'])
    update_shared_energy_storage_model_to_admm(shared_ess_data, esso_model, admm_parameters)

    # ------------------------------------------------------------------------------------------------------------------
    # ADMM -- Main cycle
    # ------------------------------------------------------------------------------------------------------------------
    convergence, num_iter = False, 1
    for iter in range(admm_parameters.num_max_iters):

        print(f'[INFO]\t - ADMM. Iter {num_iter}...')

        iter_start = time.time()

        # --------------------------------------------------------------------------------------------------------------
        # 2. Solve TSO problem
        results['tso'] = update_transmission_coordination_model_and_solve(transmission_network, tso_model,
                                                                          consensus_vars['interface']['pf']['dso'], dual_vars['pf']['tso'],
                                                                          consensus_vars['ess']['esso'], dual_vars['ess']['tso'],
                                                                          admm_parameters)

        # 2.1 Update ADMM CONSENSUS variables
        planning_problem.update_admm_consensus_variables(tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, admm_parameters)

        # 2.2 Update primal evolution
        primal_evolution.append(compute_primal_value(planning_problem, tso_model, dso_models, esso_model))

        # 2.3 STOPPING CRITERIA evaluation
        convergence = check_admm_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, admm_parameters)
        if convergence:
            break

        # --------------------------------------------------------------------------------------------------------------
        # 3. Solve DSOs problems
        results['dso'] = update_distribution_coordination_models_and_solve(distribution_networks, dso_models, consensus_vars['interface']['v'],
                                                                           consensus_vars['interface']['pf']['tso'], dual_vars['pf']['dso'],
                                                                           consensus_vars['ess']['esso'], dual_vars['ess']['dso'],
                                                                           admm_parameters)

        # 3.1 Update ADMM CONSENSUS variables
        planning_problem.update_admm_consensus_variables(tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, admm_parameters)

        # 3.2 Update primal evolution
        primal_evolution.append(compute_primal_value(planning_problem, tso_model, dso_models, esso_model))

        # 3.3 STOPPING CRITERIA evaluation
        convergence = check_admm_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, admm_parameters)
        if convergence:
            break

        # --------------------------------------------------------------------------------------------------------------
        # 4. Solve ESSO problem
        results['esso'] = update_shared_energy_storages_coordination_model_and_solve(planning_problem, esso_model,
                                                                                     consensus_vars['ess'], dual_vars['ess'],
                                                                                     admm_parameters)

        # 4.1 Update ADMM CONSENSUS variables
        planning_problem.update_admm_consensus_variables(tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, admm_parameters)

        # 4.2 Update primal evolution
        primal_evolution.append(compute_primal_value(planning_problem, tso_model, dso_models, esso_model))

        # 4.3 STOPPING CRITERIA evaluation
        convergence = check_admm_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, admm_parameters)
        if convergence:
            break

        iter_end = time.time()
        print('[INFO] \t - Iter {}: {:.2f} s'.format(num_iter, iter_end - iter_start))
        num_iter += 1

    if not convergence:
        print(f'[WARNING] ADMM did NOT converge in {admm_parameters.num_max_iters} iterations!')
    else:
        print(f'[INFO] \t - ADMM converged in {iter + 1} iterations.')

    # Used in the outer (Benders') cycle
    sensitivities = shared_ess_data.get_sensitivities(esso_model)
    optim_models = {'tso': tso_model, 'dso': dso_models, 'esso': esso_model}

    return results, sensitivities, optim_models


def create_admm_variables(planning_problem):

    num_instants = planning_problem.num_instants

    consensus_variables = {
        'interface': {
            'v': dict(),
            'pf': {'tso': dict(), 'dso': dict()}
        },
        'ess': {'tso': dict(), 'dso': dict(), 'esso': dict(), 'capacity': {'s': dict(), 'e': dict()}}
    }

    dual_variables = {
        'pf': {'tso': dict(), 'dso': dict()},
        'ess': {'tso': dict(), 'dso': dict()}
    }

    consensus_variables_prev_iter = {
        'interface': {'pf': {'tso': dict(), 'dso': dict()}},
        'ess': {'tso': dict(), 'dso': dict(), 'esso': dict()}
    }

    for dn in range(len(planning_problem.active_distribution_network_nodes)):

        node_id = planning_problem.active_distribution_network_nodes[dn]

        consensus_variables['interface']['v'][node_id] = dict()
        consensus_variables['interface']['pf']['tso'][node_id] = dict()
        consensus_variables['interface']['pf']['dso'][node_id] = dict()
        consensus_variables['ess']['tso'][node_id] = dict()
        consensus_variables['ess']['dso'][node_id] = dict()
        consensus_variables['ess']['esso'][node_id] = dict()

        dual_variables['pf']['tso'][node_id] = dict()
        dual_variables['pf']['dso'][node_id] = dict()
        dual_variables['ess']['tso'][node_id] = dict()
        dual_variables['ess']['dso'][node_id] = dict()

        consensus_variables_prev_iter['interface']['pf']['tso'][node_id] = dict()
        consensus_variables_prev_iter['interface']['pf']['dso'][node_id] = dict()
        consensus_variables_prev_iter['ess']['tso'][node_id] = dict()
        consensus_variables_prev_iter['ess']['dso'][node_id] = dict()
        consensus_variables_prev_iter['ess']['esso'][node_id] = dict()

        for year in planning_problem.years:

            consensus_variables['interface']['v'][node_id][year] = dict()
            consensus_variables['interface']['pf']['tso'][node_id][year] = dict()
            consensus_variables['interface']['pf']['dso'][node_id][year] = dict()
            consensus_variables['ess']['tso'][node_id][year] = dict()
            consensus_variables['ess']['dso'][node_id][year] = dict()
            consensus_variables['ess']['esso'][node_id][year] = dict()

            dual_variables['pf']['tso'][node_id][year] = dict()
            dual_variables['pf']['dso'][node_id][year] = dict()
            dual_variables['ess']['tso'][node_id][year] = dict()
            dual_variables['ess']['dso'][node_id][year] = dict()

            consensus_variables_prev_iter['interface']['pf']['tso'][node_id][year] = dict()
            consensus_variables_prev_iter['interface']['pf']['dso'][node_id][year] = dict()
            consensus_variables_prev_iter['ess']['tso'][node_id][year] = dict()
            consensus_variables_prev_iter['ess']['dso'][node_id][year] = dict()
            consensus_variables_prev_iter['ess']['esso'][node_id][year] = dict()

            for day in planning_problem.days:

                consensus_variables['interface']['v'][node_id][year][day] = [1.0] * num_instants
                consensus_variables['interface']['pf']['tso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['interface']['pf']['dso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['tso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['dso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['esso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

                dual_variables['pf']['tso'][node_id][year][day] = {'p': [0.0] * planning_problem.num_instants, 'q': [0.0] * num_instants}
                dual_variables['pf']['dso'][node_id][year][day] = {'p': [0.0] * planning_problem.num_instants, 'q': [0.0] * num_instants}
                dual_variables['ess']['tso'][node_id][year][day] = {'p': [0.0] * planning_problem.num_instants, 'q': [0.0] * num_instants}
                dual_variables['ess']['dso'][node_id][year][day] = {'p': [0.0] * planning_problem.num_instants, 'q': [0.0] * num_instants}

                consensus_variables_prev_iter['interface']['pf']['tso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables_prev_iter['interface']['pf']['dso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables_prev_iter['ess']['tso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables_prev_iter['ess']['dso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables_prev_iter['ess']['esso'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

    return consensus_variables, dual_variables, consensus_variables_prev_iter


def create_transmission_network_model(transmission_network, interface_v_vars, interface_pf_vars, sess_vars, candidate_solution):

    # Build model, fix candidate solution, and Run S-MPOPF model
    tso_model = transmission_network.build_model()
    transmission_network.update_model_with_candidate_solution(tso_model, candidate_solution)
    for node_id in transmission_network.active_distribution_network_nodes:
        for year in transmission_network.years:
            for day in transmission_network.days:
                node_idx = transmission_network.network[year][day].get_node_idx(node_id)
                s_base = transmission_network.network[year][day].baseMVA
                for s_m in tso_model[year][day].scenarios_market:
                    for s_o in tso_model[year][day].scenarios_operation:
                        for p in tso_model[year][day].periods:
                            pc = interface_pf_vars['dso'][node_id][year][day]['p'][p] / s_base
                            qc = interface_pf_vars['dso'][node_id][year][day]['q'][p] / s_base
                            tso_model[year][day].pc[node_idx, s_m, s_o, p].fix(pc)
                            tso_model[year][day].qc[node_idx, s_m, s_o, p].fix(qc)
                            if transmission_network.params.fl_reg:
                                tso_model[year][day].flex_p_up[node_idx, s_m, s_o, p].fix(0.0)
                                tso_model[year][day].flex_p_down[node_idx, s_m, s_o, p].fix(0.0)
    transmission_network.optimize(tso_model)

    # Get initial interface PF values
    for year in transmission_network.years:
        for day in transmission_network.days:
            s_base = transmission_network.network[year][day].baseMVA
            for dn in tso_model[year][day].active_distribution_networks:
                node_id = transmission_network.active_distribution_network_nodes[dn]
                for p in tso_model[year][day].periods:
                    v_mag = sqrt(pe.value(tso_model[year][day].expected_interface_vmag_sqr[dn, p]))
                    interface_pf_p = pe.value(tso_model[year][day].expected_interface_pf_p[dn, p]) * s_base
                    interface_pf_q = pe.value(tso_model[year][day].expected_interface_pf_q[dn, p]) * s_base
                    interface_v_vars[node_id][year][day][p] = v_mag
                    interface_pf_vars['tso'][node_id][year][day]['p'][p] = interface_pf_p
                    interface_pf_vars['tso'][node_id][year][day]['q'][p] = interface_pf_q

    # Get initial Shared ESS values
    for year in transmission_network.years:
        for day in transmission_network.days:
            s_base = transmission_network.network[year][day].baseMVA
            for dn in tso_model[year][day].active_distribution_networks:
                node_id = transmission_network.active_distribution_network_nodes[dn]
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                for p in tso_model[year][day].periods:
                    shared_ess_p = pe.value(tso_model[year][day].expected_shared_ess_p[shared_ess_idx, p]) * s_base
                    sess_vars[node_id][year][day]['p'][p] = shared_ess_p

    return tso_model


def create_distribution_networks_models(distribution_networks, interface_vars, sess_vars, candidate_solution):

    dso_models = dict()

    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]

        # Build model, fix candidate solution, and Run S-MPOPF model
        dso_model = distribution_network.build_model()
        distribution_network.update_model_with_candidate_solution(dso_model, candidate_solution)
        distribution_network.optimize(dso_model)

        # Get initial interface PF values
        for year in distribution_network.years:
            for day in distribution_network.days:
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    interface_pf_p = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                    interface_pf_q = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base
                    interface_vars[node_id][year][day]['p'][p] = interface_pf_p
                    interface_vars[node_id][year][day]['q'][p] = interface_pf_q

        # Get initial Shared ESS values
        for year in distribution_network.years:
            for day in distribution_network.days:
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    p_ess = pe.value(dso_model[year][day].expected_shared_ess_p[p]) * s_base
                    sess_vars[node_id][year][day]['p'][p] = p_ess

        dso_models[node_id] = dso_model

    return dso_models


def create_shared_energy_storage_model(shared_ess_data, candidate_solution):

    esso_model = shared_ess_data.build_subproblem()
    shared_ess_data.update_model_with_candidate_solution(esso_model, candidate_solution)
    shared_ess_data.optimize(esso_model)

    return esso_model


def update_transmission_model_to_admm(transmission_network, model, initial_interface_pf, params):

    for year in transmission_network.years:
        for day in transmission_network.days:

            init_of_value = pe.value(model[year][day].objective)
            s_base = transmission_network.network[year][day].baseMVA

            # Free Pc and Qc at the connection point with distribution networks
            for node_id in transmission_network.active_distribution_network_nodes:
                node_idx = transmission_network.network[year][day].get_node_idx(node_id)
                for s_m in model[year][day].scenarios_market:
                    for s_o in model[year][day].scenarios_operation:
                        for p in model[year][day].periods:
                            model[year][day].pc[node_idx, s_m, s_o, p].fixed = False
                            model[year][day].pc[node_idx, s_m, s_o, p].setub(None)
                            model[year][day].pc[node_idx, s_m, s_o, p].setlb(None)
                            model[year][day].qc[node_idx, s_m, s_o, p].fixed = False
                            model[year][day].qc[node_idx, s_m, s_o, p].setub(None)
                            model[year][day].qc[node_idx, s_m, s_o, p].setlb(None)

            # Add ADMM variables
            model[year][day].rho = pe.Var(domain=pe.NonNegativeReals)
            model[year][day].rho.fix(params.rho[transmission_network.name])

            # Power Flow - Consensus
            model[year][day].p_pf_req = pe.Var(model[year][day].active_distribution_networks, model[year][day].periods, domain=pe.Reals)  # Active power - requested by distribution networks
            model[year][day].q_pf_req = pe.Var(model[year][day].active_distribution_networks, model[year][day].periods, domain=pe.Reals)  # Reactive power - requested by distribution networks
            model[year][day].dual_pf_p_req = pe.Var(model[year][day].active_distribution_networks, model[year][day].periods, domain=pe.Reals)  # Dual variable - active power requested
            model[year][day].dual_pf_q_req = pe.Var(model[year][day].active_distribution_networks, model[year][day].periods, domain=pe.Reals)  # Dual variable - reactive power requested

            # Shared Energy Storage - Consensus
            model[year][day].p_ess_req = pe.Var(model[year][day].shared_energy_storages, model[year][day].periods, domain=pe.Reals)  # Shared ESS - Charging requested by DSO
            model[year][day].dual_ess_p = pe.Var(model[year][day].shared_energy_storages, model[year][day].periods, domain=pe.Reals)  # Dual variable - Shared ESS active power

            # Objective function - augmented Lagrangian
            obj = model[year][day].objective.expr / abs(init_of_value)
            for dn in model[year][day].active_distribution_networks:
                node_id = transmission_network.active_distribution_network_nodes[dn]
                for p in model[year][day].periods:
                    init_p = initial_interface_pf['dso'][node_id][year][day]['p'][p] / s_base
                    init_q = initial_interface_pf['dso'][node_id][year][day]['q'][p] / s_base
                    constraint_p_req = (model[year][day].expected_interface_pf_p[dn, p] - model[year][day].p_pf_req[dn, p]) / abs(init_p)
                    constraint_q_req = (model[year][day].expected_interface_pf_q[dn, p] - model[year][day].q_pf_req[dn, p]) / abs(init_q)
                    obj += model[year][day].dual_pf_p_req[dn, p] * constraint_p_req
                    obj += model[year][day].dual_pf_q_req[dn, p] * constraint_q_req
                    obj += (model[year][day].rho / 2) * constraint_p_req ** 2
                    obj += (model[year][day].rho / 2) * constraint_q_req ** 2

            for e in model[year][day].active_distribution_networks:
                rating = transmission_network.network[year][day].shared_energy_storages[e].s
                if rating == 0.0:
                    rating = 1.00       # Do not balance residuals
                for p in model[year][day].periods:
                    constraint_ess_p = (model[year][day].expected_shared_ess_p[e, p] - model[year][day].p_ess_req[e, p]) / (2 * rating)
                    obj += model[year][day].dual_ess_p[e, p] * constraint_ess_p
                    obj += (model[year][day].rho / 2) * constraint_ess_p ** 2

            model[year][day].objective.expr = obj


def update_distribution_models_to_admm(distribution_networks, models, initial_interface_pf, params):

    for node_id in distribution_networks:

        dso_model = models[node_id]
        distribution_network = distribution_networks[node_id]

        # Free voltage at the connection point with the transmission network
        # Free Pg and Qg at the connection point with the transmission network
        for year in distribution_network.years:
            for day in distribution_network.days:

                init_of_value = pe.value(dso_model[year][day].objective)
                rating = distribution_network.network[year][day].shared_energy_storages[0].s
                if rating == 0.0:
                    rating = 1.00                # Do not balance residuals
                    init_of_value = 1.00

                ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                ref_node_idx = distribution_network.network[year][day].get_node_idx(ref_node_id)
                ref_gen_idx = distribution_network.network[year][day].get_reference_gen_idx()
                for s_m in dso_model[year][day].scenarios_market:
                    for s_o in dso_model[year][day].scenarios_operation:
                        for p in dso_model[year][day].periods:
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].fixed = False
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].setub(None)
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].setlb(None)

                            dso_model[year][day].pg[ref_gen_idx, s_m, s_o, p].fixed = False
                            dso_model[year][day].pg[ref_gen_idx, s_m, s_o, p].setub(None)
                            dso_model[year][day].pg[ref_gen_idx, s_m, s_o, p].setlb(None)
                            dso_model[year][day].qg[ref_gen_idx, s_m, s_o, p].fixed = False
                            dso_model[year][day].qg[ref_gen_idx, s_m, s_o, p].setub(None)
                            dso_model[year][day].qg[ref_gen_idx, s_m, s_o, p].setlb(None)

                # Add ADMM variables
                dso_model[year][day].rho = pe.Var(domain=pe.NonNegativeReals)
                dso_model[year][day].rho.fix(params.rho[distribution_network.network[year][day].name])

                dso_model[year][day].p_pf_req = pe.Var(dso_model[year][day].periods, domain=pe.Reals)    # Active power - requested by transmission network
                dso_model[year][day].q_pf_req = pe.Var(dso_model[year][day].periods, domain=pe.Reals)    # Reactive power - requested by transmission network
                dso_model[year][day].dual_pf_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals)   # Dual variable - active power
                dso_model[year][day].dual_pf_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals)   # Dual variable - reactive power

                dso_model[year][day].p_ess_req = pe.Var(dso_model[year][day].periods, domain=pe.Reals)   # Shared ESS - Charging requested by TSO
                dso_model[year][day].dual_ess_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals)  # Dual variable - Shared ESS active power

                # Objective function - augmented Lagrangian
                obj = dso_model[year][day].objective.expr / max(abs(init_of_value), 1.00)

                # Augmented Lagrangian -- Interface power flow (residual balancing)
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    init_p = initial_interface_pf[node_id][year][day]['p'][p] / s_base
                    init_q = initial_interface_pf[node_id][year][day]['q'][p] / s_base
                    constraint_p_req = (dso_model[year][day].expected_interface_pf_p[p] - dso_model[year][day].p_pf_req[p]) / abs(init_p)
                    constraint_q_req = (dso_model[year][day].expected_interface_pf_q[p] - dso_model[year][day].q_pf_req[p]) / abs(init_q)
                    obj += (dso_model[year][day].dual_pf_p[p]) * (constraint_p_req)
                    obj += (dso_model[year][day].dual_pf_q[p]) * (constraint_q_req)
                    obj += (dso_model[year][day].rho / 2) * (constraint_p_req) ** 2
                    obj += (dso_model[year][day].rho / 2) * (constraint_q_req) ** 2

                # Augmented Lagrangian -- Shared ESS (residual balancing)
                for p in dso_model[year][day].periods:
                    constraint_ess_p = (dso_model[year][day].expected_shared_ess_p[p] - dso_model[year][day].p_ess_req[p]) / (2 * rating)
                    obj += dso_model[year][day].dual_ess_p[p] * (constraint_ess_p)
                    obj += (dso_model[year][day].rho / 2) * (constraint_ess_p) ** 2

                dso_model[year][day].objective.expr = obj


def update_shared_energy_storage_model_to_admm(shared_ess_data, model, params):

    repr_years = [year for year in shared_ess_data.years]

    # Add ADMM variables
    model.rho = pe.Var(domain=pe.NonNegativeReals)
    model.rho.fix(params.rho['ESSO'])

    # Active and Reactive power requested by TSO and DSOs
    model.p_req_transm = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.Reals)   # Active power - transmission network
    model.p_req_distr = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.Reals)    # Active power - distribution networks
    model.dual_p_transm = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.Reals)  # Dual variable - active power - transmission network
    model.dual_p_distr = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.Reals)   # Dual variable - active power - distribution networks

    # Objective function - augmented Lagrangian
    init_of_value = pe.value(model.objective)
    obj = model.objective.expr / abs(init_of_value)
    for e in model.energy_storages:
        for y in model.years:
            year = repr_years[y]
            rating_s = shared_ess_data.shared_energy_storages[year][e].s
            if rating_s == 0.0:
                rating_s = 1.00     # Do not balance residuals
            for d in model.days:
                for p in model.periods:
                    p_ess = model.es_expected_p[e, y, d, p]
                    constraint_p_transm = (p_ess - model.p_req_transm[e, y, d, p]) / (2 * rating_s)
                    constraint_p_distr = (p_ess - model.p_req_distr[e, y, d, p]) / (2 * rating_s)
                    obj += model.dual_p_transm[e, y, d, p] * (constraint_p_transm)
                    obj += model.dual_p_distr[e, y, d, p] * (constraint_p_distr)
                    obj += (model.rho / 2) * (constraint_p_transm) ** 2
                    obj += (model.rho / 2) * (constraint_p_distr) ** 2

    model.objective.expr = obj

    return model


def update_transmission_coordination_model_and_solve(transmission_network, model, pf_req, dual_pf, ess_req, dual_ess, params):

    print('[INFO] \t\t - Updating transmission network...')

    for year in transmission_network.years:
        for day in transmission_network.days:

            s_base = transmission_network.network[year][day].baseMVA

            # Update Rho parameter
            model[year][day].rho.fix(params.rho[transmission_network.name])

            for dn in model[year][day].active_distribution_networks:

                node_id = transmission_network.active_distribution_network_nodes[dn]

                # Update interface PF power requests
                for p in model[year][day].periods:
                    model[year][day].dual_pf_p_req[dn, p].fix(dual_pf[node_id][year][day]['p'][p] / s_base)
                    model[year][day].dual_pf_q_req[dn, p].fix(dual_pf[node_id][year][day]['q'][p] / s_base)
                    model[year][day].p_pf_req[dn, p].fix(pf_req[node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_pf_req[dn, p].fix(pf_req[node_id][year][day]['q'][p] / s_base)

                # Update shared ESS capacity and power requests
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                for p in model[year][day].periods:
                    model[year][day].dual_ess_p[shared_ess_idx, p].fix(dual_ess[node_id][year][day]['p'][p] / s_base)
                    model[year][day].p_ess_req[shared_ess_idx, p].fix(ess_req[node_id][year][day]['p'][p] / s_base)

    # Solve!
    res = transmission_network.optimize(model, from_warm_start=True)
    for year in transmission_network.years:
        for day in transmission_network.days:
            if res[year][day].solver.status == po.SolverStatus.error:
                print(f'[ERROR] Network {model[year][day].name} did not converge!')
                exit(ERROR_NETWORK_OPTIMIZATION)
    return res


def update_distribution_coordination_models_and_solve(distribution_networks, models, interface_vmag, pf_req, dual_pf, ess_req, dual_ess, params):

    print('[INFO] \t\t - Updating distribution networks:')
    res = dict()

    for node_id in distribution_networks:

        model = models[node_id]
        distribution_network = distribution_networks[node_id]
        rho = params.rho[distribution_network.name]

        print('[INFO] \t\t\t - Updating active distribution network connected to node {}...'.format(node_id))

        for year in distribution_network.years:
            for day in distribution_network.days:

                s_base = distribution_network.network[year][day].baseMVA
                ref_node_id = distribution_network.network[year][day].get_reference_node_id()

                model[year][day].rho.fix(rho)

                # Update VOLTAGE variables at connection point
                for p in model[year][day].periods:
                    model[year][day].expected_interface_vmag_sqr[p].fix(interface_vmag[node_id][year][day][p]**2)

                # Update POWER FLOW variables at connection point
                for p in model[year][day].periods:
                    model[year][day].dual_pf_p[p].fix(dual_pf[node_id][year][day]['p'][p] / s_base)
                    model[year][day].dual_pf_q[p].fix(dual_pf[node_id][year][day]['q'][p] / s_base)
                    model[year][day].p_pf_req[p].fix(pf_req[node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_pf_req[p].fix(pf_req[node_id][year][day]['q'][p] / s_base)

                # Update SHARED ENERGY STORAGE variables (if existent)
                shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                for p in model[year][day].periods:
                    model[year][day].dual_ess_p[p].fix(dual_ess[node_id][year][day]['p'][p] / s_base)
                    model[year][day].p_ess_req[p].fix(ess_req[node_id][year][day]['p'][p] / s_base)

        # Solve!
        res[node_id] = distribution_network.optimize(model, from_warm_start=True)
        for year in distribution_network.years:
            for day in distribution_network.days:
                if res[node_id][year][day].solver.status != po.SolverStatus.ok:
                    print(f'[WARNING] Network {model[year][day].name} did not converge!')
                    #exit(ERROR_NETWORK_OPTIMIZATION)
    return res


def update_shared_energy_storages_coordination_model_and_solve(planning_problem, model, ess_req, dual_ess, params):

    print('[INFO] \t\t - Updating Shared ESS...')
    shared_ess_data = planning_problem.shared_ess_data
    days = [day for day in planning_problem.days]
    years = [year for year in planning_problem.years]

    model.rho.fix(params.rho['ESSO'])

    for e in model.energy_storages:
        for y in model.years:
            year = years[y]
            node_id = shared_ess_data.shared_energy_storages[year][e].bus
            for d in model.days:
                day = days[d]
                for p in model.periods:
                    model.dual_p_transm[e, y, d, p].fix(dual_ess['tso'][node_id][year][day]['p'][p])
                    model.dual_p_distr[e, y, d, p].fix(dual_ess['dso'][node_id][year][day]['p'][p])
                    model.p_req_transm[e, y, d, p].fix(ess_req['tso'][node_id][year][day]['p'][p])
                    model.p_req_distr[e, y, d, p].fix(ess_req['dso'][node_id][year][day]['p'][p])

    # Solve!
    res = shared_ess_data.optimize(model, from_warm_start=True)
    if res.solver.status != po.SolverStatus.ok:
        print('[WARNING] Shared ESS operational planning did not converge!')

    return res


def _update_admm_consensus_variables(planning_problem, tso_model, dso_models, esso_model, consensus_vars, dual_vars, consensus_vars_prev_iter, params):
    _update_previous_consensus_variables(planning_problem, consensus_vars, consensus_vars_prev_iter)
    _update_interface_power_flow_variables(planning_problem, tso_model, dso_models, consensus_vars['interface'], dual_vars['pf'], params)
    _update_shared_energy_storage_variables(planning_problem, tso_model, dso_models, esso_model, consensus_vars['ess'], dual_vars['ess'], params)


def _update_previous_consensus_variables(planning_problem, consensus_vars, consensus_vars_prev_iter):
    for dn in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[dn]
        for year in planning_problem.years:
            for day in planning_problem.days:
                for p in range(planning_problem.num_instants):
                    consensus_vars_prev_iter['interface']['pf']['tso'][node_id][year][day]['p'][p] = copy(consensus_vars['interface']['pf']['tso'][node_id][year][day]['p'][p])
                    consensus_vars_prev_iter['interface']['pf']['tso'][node_id][year][day]['q'][p] = copy(consensus_vars['interface']['pf']['tso'][node_id][year][day]['q'][p])
                    consensus_vars_prev_iter['interface']['pf']['dso'][node_id][year][day]['p'][p] = copy(consensus_vars['interface']['pf']['dso'][node_id][year][day]['p'][p])
                    consensus_vars_prev_iter['interface']['pf']['dso'][node_id][year][day]['q'][p] = copy(consensus_vars['interface']['pf']['dso'][node_id][year][day]['q'][p])
                    consensus_vars_prev_iter['ess']['tso'][node_id][year][day]['p'][p] = copy(consensus_vars['ess']['tso'][node_id][year][day]['p'][p])
                    consensus_vars_prev_iter['ess']['dso'][node_id][year][day]['p'][p] = copy(consensus_vars['ess']['dso'][node_id][year][day]['p'][p])
                    consensus_vars_prev_iter['ess']['esso'][node_id][year][day]['p'][p] = copy(consensus_vars['ess']['esso'][node_id][year][day]['p'][p])


def _update_interface_power_flow_variables(planning_problem, tso_model, dso_models, interface_vars, dual_vars, params):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    # Transmission network - Update Vmag and PF at the TN-DN interface
    for dn in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[dn]
        for year in planning_problem.years:
            for day in planning_problem.days:
                s_base = planning_problem.transmission_network.network[year][day].baseMVA
                for p in tso_model[year][day].periods:
                    interface_vars['v'][node_id][year][day][p] = sqrt(pe.value(tso_model[year][day].expected_interface_vmag_sqr[dn, p]))
                    interface_vars['pf']['tso'][node_id][year][day]['p'][p] = pe.value(tso_model[year][day].expected_interface_pf_p[dn, p]) * s_base
                    interface_vars['pf']['tso'][node_id][year][day]['q'][p] = pe.value(tso_model[year][day].expected_interface_pf_q[dn, p]) * s_base

    # Distribution Network - Update PF at the TN-DN interface
    for node_id in distribution_networks:
        distribution_network = distribution_networks[node_id]
        dso_model = dso_models[node_id]
        for year in planning_problem.years:
            for day in planning_problem.days:
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    interface_vars['pf']['dso'][node_id][year][day]['p'][p] = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                    interface_vars['pf']['dso'][node_id][year][day]['q'][p] = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base

    # Update Lambdas
    for node_id in distribution_networks:
        distribution_network = distribution_networks[node_id]
        for year in planning_problem.years:
            for day in planning_problem.days:
                for p in range(planning_problem.num_instants):

                    error_p_pf = interface_vars['pf']['tso'][node_id][year][day]['p'][p] - interface_vars['pf']['dso'][node_id][year][day]['p'][p]
                    error_q_pf = interface_vars['pf']['tso'][node_id][year][day]['q'][p] - interface_vars['pf']['dso'][node_id][year][day]['q'][p]

                    dual_vars['tso'][node_id][year][day]['p'][p] += params.rho[transmission_network.name] * (error_p_pf)
                    dual_vars['tso'][node_id][year][day]['q'][p] += params.rho[transmission_network.name] * (error_q_pf)
                    dual_vars['dso'][node_id][year][day]['p'][p] += params.rho[distribution_network.name] * (-error_p_pf)
                    dual_vars['dso'][node_id][year][day]['q'][p] += params.rho[distribution_network.name] * (-error_q_pf)

                '''
                print(f"Ptso[{node_id},{year},{day}] = {interface_vars['pf']['tso'][node_id][year][day]['p']}")
                print(f"Pdso[{node_id},{year},{day}] = {interface_vars['pf']['dso'][node_id][year][day]['p']}")
                print(f"Qtso[{node_id},{year},{day}] = {interface_vars['pf']['tso'][node_id][year][day]['q']}")
                print(f"Qdso[{node_id},{year},{day}] = {interface_vars['pf']['dso'][node_id][year][day]['q']}")
                '''


def _update_shared_energy_storage_variables(planning_problem, tso_model, dso_models, sess_model, shared_ess_vars, dual_vars, params):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data
    repr_days = [day for day in planning_problem.days]
    repr_years = [year for year in planning_problem.years]

    for node_id in distribution_networks:

        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]

        # Shared Energy Storage - Power requested by ESSO, and Capacity available
        for y in sess_model.years:
            year = repr_years[y]
            shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)
            for d in sess_model.days:
                day = repr_days[d]
                shared_ess_vars['esso'][node_id][year][day]['p'] = [0.0 for _ in range(planning_problem.num_instants)]
                for p in sess_model.periods:
                    shared_ess_vars['esso'][node_id][year][day]['p'][p] = pe.value(sess_model.es_expected_p[shared_ess_idx, y, d, p])

        # Shared Energy Storage - Power requested by TSO
        for y in range(len(planning_problem.years)):
            year = repr_years[y]
            for d in range(len(repr_days)):
                day = repr_days[d]
                s_base = transmission_network.network[year][day].baseMVA
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                shared_ess_vars['tso'][node_id][year][day]['p'] = [0.0 for _ in range(planning_problem.num_instants)]
                for p in tso_model[year][day].periods:
                    shared_ess_vars['tso'][node_id][year][day]['p'][p] = pe.value(tso_model[year][day].expected_shared_ess_p[shared_ess_idx, p]) * s_base

        # Shared Energy Storage - Power requested by DSO
        for y in range(len(planning_problem.years)):
            year = repr_years[y]
            for d in range(len(repr_days)):
                day = repr_days[d]
                s_base = distribution_network.network[year][day].baseMVA
                shared_ess_vars['dso'][node_id][year][day]['p'] = [0.0 for _ in range(planning_problem.num_instants)]
                for p in dso_model[year][day].periods:
                    shared_ess_vars['dso'][node_id][year][day]['p'][p] = pe.value(dso_model[year][day].expected_shared_ess_p[p]) * s_base

        '''
        for year in planning_problem.years:
            for day in planning_problem.days:
                print(f"Preq, TN, Node {node_id}, {year}, {day} = {shared_ess_vars['tso'][node_id][year][day]['p']}")
                print(f"Preq, DN, Node {node_id}, {year}, {day} = {shared_ess_vars['dso'][node_id][year][day]['p']}")
                print(f"Preq, ESS, Node {node_id}, {year}, {day} = {shared_ess_vars['esso'][node_id][year][day]['p']}")
        '''

        # Update dual variables Shared ESS
        for year in planning_problem.years:
            for day in planning_problem.days:
                for t in range(planning_problem.num_instants):
                    error_p_ess_transm = shared_ess_vars['tso'][node_id][year][day]['p'][t] - shared_ess_vars['esso'][node_id][year][day]['p'][t]
                    error_p_ess_distr = shared_ess_vars['dso'][node_id][year][day]['p'][t] - shared_ess_vars['esso'][node_id][year][day]['p'][t]
                    dual_vars['tso'][node_id][year][day]['p'][t] += params.rho['ESSO'] * (error_p_ess_transm)
                    dual_vars['dso'][node_id][year][day]['p'][t] += params.rho['ESSO'] * (error_p_ess_distr)


def compute_primal_value(planning_problem, tso_model, dso_models, esso_model):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data

    primal_value = 0.0
    primal_value += transmission_network.compute_primal_value(tso_model, transmission_network.params)
    for node_id in distribution_networks:
        primal_value +=distribution_networks[node_id].compute_primal_value(dso_models[node_id], distribution_networks[node_id].params)
    if esso_model:
        primal_value += shared_ess_data.compute_primal_value(esso_model)

    return primal_value


def check_admm_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, params):
    if consensus_convergence(planning_problem, consensus_vars, params):
        if stationary_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, params):
            return True
    return False


def consensus_convergence(planning_problem, consensus_vars, params):

    interface_vars = consensus_vars['interface']['pf']
    shared_ess_vars = consensus_vars['ess']
    sum_sqr = 0.0
    num_elems = 0

    # Interface Power Flow
    for node_id in planning_problem.active_distribution_network_nodes:
        for year in planning_problem.years:
            for day in planning_problem.days:
                for p in range(planning_problem.num_instants):
                    sum_sqr += (interface_vars['tso'][node_id][year][day]['p'][p] - interface_vars['dso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += (interface_vars['tso'][node_id][year][day]['q'][p] - interface_vars['dso'][node_id][year][day]['q'][p]) ** 2
                    num_elems += 2

    # Shared Energy Storage
    for node_id in planning_problem.active_distribution_network_nodes:
        shared_ess_idx = planning_problem.shared_ess_data.get_shared_energy_storage_idx(node_id)
        for year in planning_problem.years:
            if planning_problem.shared_ess_data.shared_energy_storages[year][shared_ess_idx].s == 0:
                continue
            for day in planning_problem.days:
                for p in range(planning_problem.num_instants):
                    sum_sqr += (shared_ess_vars['tso'][node_id][year][day]['p'][p] - shared_ess_vars['esso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += (shared_ess_vars['dso'][node_id][year][day]['p'][p] - shared_ess_vars['esso'][node_id][year][day]['p'][p]) ** 2
                    num_elems += 2

    sum_total = sqrt(sum_sqr)
    if sum_total > params.tol * sqrt(num_elems) and not isclose(sum_total, params.tol * sqrt(num_elems), rel_tol=ADMM_CONVERGENCE_REL_TOL, abs_tol=params.tol):
        print('[INFO]\t\t - Convergence primal failed. {:.3f} > {:.3f}'.format(sum_total, params.tol * sqrt(num_elems)))
        return False

    #print('Convergence primal ok. {} <= {}'.format(sum_total, params.tol * sqrt(num_elems)))
    return True


def stationary_convergence(planning_problem, consensus_vars, consensus_vars_prev_iter, params):

    rho_esso = params.rho['ESSO']
    rho_tso = params.rho[planning_problem.transmission_network.name]
    interface_vars = consensus_vars['interface']['pf']
    shared_ess_vars = consensus_vars['ess']
    interface_vars_prev_iter = consensus_vars_prev_iter['interface']['pf']
    shared_ess_vars_prev_iter = consensus_vars_prev_iter['ess']
    sum_sqr = 0.0
    num_elems = 0

    # Interface Power Flow
    for node_id in planning_problem.distribution_networks:
        rho_dso = params.rho[planning_problem.distribution_networks[node_id].name]
        for year in planning_problem.years:
            for day in planning_problem.days:
                for p in range(planning_problem.num_instants):
                    sum_sqr += rho_tso * (interface_vars['tso'][node_id][year][day]['p'][p] - interface_vars_prev_iter['tso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += rho_tso * (interface_vars['tso'][node_id][year][day]['q'][p] - interface_vars_prev_iter['tso'][node_id][year][day]['q'][p]) ** 2
                    sum_sqr += rho_dso * (interface_vars['dso'][node_id][year][day]['p'][p] - interface_vars_prev_iter['dso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += rho_dso * (interface_vars['dso'][node_id][year][day]['q'][p] - interface_vars_prev_iter['dso'][node_id][year][day]['q'][p]) ** 2
                    num_elems += 4

    # Shared Energy Storage
    for node_id in planning_problem.distribution_networks:
        distribution_network = planning_problem.distribution_networks[node_id]
        for year in planning_problem.years:
            for day in planning_problem.days:
                rho_dso = params.rho[distribution_network.network[year][day].name]
                for p in range(planning_problem.num_instants):
                    sum_sqr += rho_tso * (shared_ess_vars['tso'][node_id][year][day]['p'][p] - shared_ess_vars_prev_iter['tso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += rho_dso * (shared_ess_vars['dso'][node_id][year][day]['p'][p] - shared_ess_vars_prev_iter['dso'][node_id][year][day]['p'][p]) ** 2
                    sum_sqr += rho_esso * (shared_ess_vars['esso'][node_id][year][day]['p'][p] - shared_ess_vars_prev_iter['esso'][node_id][year][day]['p'][p]) ** 2
                    num_elems += 3

    sum_total = sqrt(sum_sqr)
    if sum_total > params.tol * sqrt(num_elems) and not isclose(sum_total, params.tol * sqrt(num_elems), rel_tol=ADMM_CONVERGENCE_REL_TOL, abs_tol=params.tol):
        print('[INFO]\t\t - Convergence dual failed. {:.3f} > {:.3f}'.format(sum_total, params.tol * sqrt(num_elems)))
        return False

    #print('Convergence dual ok. {} <= {}'.format(sum_total, params.tol * sqrt(num_elems)))
    return True


# ======================================================================================================================
#  OPERATIONAL PLANNING WITHOUT COORDINATION functions
# ======================================================================================================================
def _run_operational_planning_without_coordination(planning_problem):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    results = {'tso': dict(), 'dso': dict(), 'esso': dict()}

    # Do not consider flexible resources
    transmission_network.params.fl_reg = False
    transmission_network.params.es_reg = False
    transmission_network.params.transf_reg = False
    transmission_network.params.rg_curt = True
    transmission_network.params.l_curt = True
    transmission_network.params.slack_line_limits = True
    transmission_network.params.slack_voltage_limits = True
    for node_id in distribution_networks:
        distribution_network = distribution_networks[node_id]
        distribution_network.params.fl_reg = False
        distribution_network.params.es_reg = False
        distribution_network.params.transf_reg = False
        distribution_network.params.rg_curt = True
        distribution_network.params.l_curt = True
        distribution_network.params.slack_line_limits = True
        distribution_network.params.slack_voltage_limits = True

    # Shared ESS candidate solution (no hared ESS)
    candidate_solution = dict()
    for e in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[e]
        candidate_solution[node_id] = dict()
        for year in planning_problem.years:
            candidate_solution[node_id][year] = dict()
            candidate_solution[node_id][year]['s'] = 0.00
            candidate_solution[node_id][year]['e'] = 0.00

    # Create interface PF variables
    interface_pf = create_interface_power_flow_variables(planning_problem)

    # Create DSOs' Operational Planning models
    dso_models = dict()
    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]
        results['dso'][node_id] = dict()

        # Build model, fix candidate solution, and Run S-MPOPF model
        dso_model = distribution_network.build_model()
        distribution_network.update_model_with_candidate_solution(dso_model, candidate_solution)
        results['dso'][node_id] = distribution_network.optimize(dso_model)

        # Get initial interface PF values
        for year in distribution_network.years:
            for day in distribution_network.days:
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    interface_pf[node_id][year][day]['p'][p] = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                    interface_pf[node_id][year][day]['q'][p] = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base

        dso_models[node_id] = dso_model

    # Create TSO Operational Planning model
    tso_model = transmission_network.build_model()
    transmission_network.update_model_with_candidate_solution(tso_model, candidate_solution)
    for node_id in transmission_network.active_distribution_network_nodes:
        for year in transmission_network.years:
            for day in transmission_network.days:

                node_idx = transmission_network.network[year][day].get_node_idx(node_id)
                s_base = transmission_network.network[year][day].baseMVA

                # - Fix expected interface PF
                pc = interface_pf[node_id][year][day]['p'][p] / s_base
                qc = interface_pf[node_id][year][day]['q'][p] / s_base
                for s_m in tso_model[year][day].scenarios_market:
                    for s_o in tso_model[year][day].scenarios_operation:
                        for p in tso_model[year][day].periods:
                            tso_model[year][day].pc[node_idx, s_m, s_o, p].fix(pc)
                            tso_model[year][day].qc[node_idx, s_m, s_o, p].fix(qc)
                            if transmission_network.params.fl_reg:
                                tso_model[year][day].flex_p_up[node_idx, s_m, s_o, p].fix(0.0)
                                tso_model[year][day].flex_p_down[node_idx, s_m, s_o, p].fix(0.0)

    results['tso'] = transmission_network.optimize(tso_model)

    # Write results to xlsx file
    planning_problem.write_operational_planning_results_without_coordination_to_excel(tso_model, dso_models, results)

    return


def create_interface_power_flow_variables(planning_problem):
    consensus_vars, _, _ = create_admm_variables(planning_problem)
    return consensus_vars['interface']['pf']['dso']


# ======================================================================================================================
#  PLANNING PROBLEM read functions
# ======================================================================================================================
def _read_planning_problem(planning_problem):

    try:
        print(f'[INFO] Reading PROBLEM SPECIFICATION from file {planning_problem.filename} ...')

        # Create results folder
        if not os.path.exists(planning_problem.results_dir):
            os.makedirs(planning_problem.results_dir)

        # Create plots (results) folder
        if not os.path.exists(planning_problem.plots_dir):
            os.makedirs(planning_problem.plots_dir)

        # Create diagrams folder
        if not os.path.exists(planning_problem.diagrams_dir):
            os.makedirs(planning_problem.diagrams_dir)

        filename = os.path.join(planning_problem.data_dir, planning_problem.filename)
        with open(filename, 'r') as file:

            lines = file.read().splitlines()

            for i in range(len(lines)):

                tokens = lines[i].split(':')

                if tokens[0] == 'Years':
                    num_years = int(tokens[1])
                    for j in range(num_years):
                        year_tokens = lines[i + j + 1].split('\t')
                        year_name = year_tokens[0]
                        num_years = int(year_tokens[1])
                        if num_years > 0:
                            planning_problem.years[year_name] = num_years

                elif tokens[0] == 'Days':
                    num_days = int(tokens[1])
                    for j in range(num_days):
                        day_tokens = lines[i + j + 1].split('\t')
                        day_name = day_tokens[0]
                        num_days = int(day_tokens[1])
                        if num_days > 0:
                            planning_problem.days[day_name] = num_days

                elif tokens[0] == 'NumInstants':
                    planning_problem.num_instants = int(tokens[1])

                elif tokens[0] == 'Discount Factor':
                    planning_problem.discount_factor = float(tokens[1])

                elif tokens[0] == 'Market Data':
                    i = i + 1
                    print('[INFO] Reading MARKET DATA from file(s)...')
                    planning_problem.market_data_file = lines[i].strip()
                    planning_problem.read_market_data_from_file()

                elif tokens[0] == 'Distribution Networks':

                    print('[INFO] Reading DISTRIBUTION NETWORK DATA from file(s)...')

                    num_dist_networks = int(tokens[1])

                    for _ in range(num_dist_networks):

                        i = i + 1
                        dn_tokens = lines[i].split('\t')

                        network_name = dn_tokens[0]             # Network filename
                        params_file = dn_tokens[1]              # Params filename
                        connection_nodeid = int(dn_tokens[2])   # Connection node ID

                        distribution_network = NetworkPlanning()
                        distribution_network.name = network_name
                        distribution_network.is_transmission = False
                        distribution_network.data_dir = planning_problem.data_dir
                        distribution_network.results_dir = planning_problem.results_dir
                        distribution_network.plots_dir = planning_problem.plots_dir
                        distribution_network.diagrams_dir = planning_problem.diagrams_dir
                        distribution_network.years = planning_problem.years
                        distribution_network.days = planning_problem.days
                        distribution_network.num_instants = planning_problem.num_instants
                        distribution_network.discount_factor = planning_problem.discount_factor
                        distribution_network.prob_market_scenarios = planning_problem.prob_market_scenarios
                        distribution_network.cost_energy_p = planning_problem.cost_energy_p
                        distribution_network.params_file = params_file
                        distribution_network.read_network_parameters()
                        if distribution_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                            distribution_network.prob_market_scenarios = [1.00]
                        distribution_network.read_network_planning_data()
                        distribution_network.tn_connection_nodeid = connection_nodeid

                        planning_problem.distribution_networks[connection_nodeid] = distribution_network
                    planning_problem.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]

                elif tokens[0] == 'Transmission Network':

                    print('[INFO] Reading TRANSMISSION NETWORK DATA from file(s)...')

                    i = i + 1
                    tn_tokens = lines[i].split('\t')

                    network_name = tn_tokens[0]     # Network filename
                    params_file = tn_tokens[1]      # Params filename

                    transmission_network = NetworkPlanning()
                    transmission_network.name = network_name
                    transmission_network.is_transmission = True
                    transmission_network.data_dir = planning_problem.data_dir
                    transmission_network.results_dir = planning_problem.results_dir
                    transmission_network.plots_dir = planning_problem.plots_dir
                    transmission_network.diagrams_dir = planning_problem.diagrams_dir
                    transmission_network.years = planning_problem.years
                    transmission_network.days = planning_problem.days
                    transmission_network.num_instants = planning_problem.num_instants
                    transmission_network.discount_factor = planning_problem.discount_factor
                    transmission_network.prob_market_scenarios = planning_problem.prob_market_scenarios
                    transmission_network.cost_energy_p = planning_problem.cost_energy_p
                    transmission_network.params_file = params_file
                    transmission_network.read_network_parameters()
                    if transmission_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                        transmission_network.prob_market_scenarios = [1.00]
                    transmission_network.read_network_planning_data()

                    transmission_network.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]
                    for year in transmission_network.years:
                        for day in transmission_network.days:
                            transmission_network.network[year][day].active_distribution_network_nodes = transmission_network.active_distribution_network_nodes
                    planning_problem.transmission_network = transmission_network

                elif tokens[0] == 'Shared Energy Storage':

                    print('[INFO] Reading SHARED ESS DATA from file(s)...')

                    i = i + 1
                    ess_tokens = lines[i].split('\t')  # Shared ESS filename, params filename

                    params_file = ess_tokens[0]
                    data_file = ess_tokens[1]

                    shared_ess_data = SharedEnergyStorageData()
                    shared_ess_data.name = planning_problem.name
                    shared_ess_data.data_dir = planning_problem.data_dir
                    shared_ess_data.results_dir = planning_problem.results_dir
                    shared_ess_data.plots_dir = planning_problem.plots_dir
                    shared_ess_data.years = planning_problem.years
                    shared_ess_data.days = planning_problem.days
                    shared_ess_data.num_instants = planning_problem.num_instants
                    shared_ess_data.discount_factor = planning_problem.discount_factor
                    shared_ess_data.prob_market_scenarios = planning_problem.prob_market_scenarios
                    shared_ess_data.cost_energy_p = planning_problem.cost_energy_p
                    shared_ess_data.cost_secondary_reserve = planning_problem.cost_secondary_reserve
                    shared_ess_data.cost_tertiary_reserve_up = planning_problem.cost_tertiary_reserve_up
                    shared_ess_data.cost_tertiary_reserve_down = planning_problem.cost_tertiary_reserve_down
                    shared_ess_data.params_file = params_file
                    shared_ess_data.read_parameters_from_file()
                    shared_ess_data.create_shared_energy_storages(planning_problem)
                    shared_ess_data.data_file = data_file
                    shared_ess_data.read_shared_energy_storage_data_from_file()
                    shared_ess_data.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]

                    planning_problem.shared_ess_data = shared_ess_data

                elif tokens[0] == 'Planning Parameters':
                    i = i + 1
                    planning_problem.params_file = lines[i].strip()
                    planning_problem.read_planning_parameters_from_file()

    except:
        print('[ERROR] Reading planning problem description! Exiting...')
        file.close()
        exit(ERROR_SPECIFICATION_FILE)
    finally:
        file.close()

    _add_shared_energy_storage_to_transmission_network(planning_problem)
    _add_shared_energy_storage_to_distribution_network(planning_problem)


# ======================================================================================================================
#  MARKET DATA read functions
# ======================================================================================================================
def _read_market_data_from_file(planning_problem):

    try:
        for year in planning_problem.years:
            filename = os.path.join(planning_problem.data_dir, 'Market Data', f'{planning_problem.market_data_file}_{year}.xlsx')
            num_scenarios, prob_scenarios = _get_market_scenarios_info_from_excel_file(filename, 'Scenarios')
            planning_problem.prob_market_scenarios = prob_scenarios
            planning_problem.cost_energy_p[year] = dict()
            planning_problem.cost_secondary_reserve[year] = dict()
            planning_problem.cost_tertiary_reserve_up[year] = dict()
            planning_problem.cost_tertiary_reserve_down[year] = dict()
            for day in planning_problem.days:
                planning_problem.cost_energy_p[year][day] = _get_market_costs_from_excel_file(filename, f'Cp, {day}', num_scenarios)
                planning_problem.cost_secondary_reserve[year][day] = _get_market_costs_from_excel_file(filename, f'Csr, {day}', num_scenarios)
                planning_problem.cost_tertiary_reserve_up[year][day] = _get_market_costs_from_excel_file(filename, f'Ctr_up, {day}', num_scenarios)
                planning_problem.cost_tertiary_reserve_down[year][day] = _get_market_costs_from_excel_file(filename, f'Ctr_down, {day}', num_scenarios)
    except:
        print(f'[ERROR] Reading market data from file(s). Exiting...')
        exit(ERROR_SPECIFICATION_FILE)


def _get_market_scenarios_info_from_excel_file(filename, sheet_name):

    num_scenarios = 0
    prob_scenarios = list()

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        if is_int(df.iloc[0, 1]):
            num_scenarios = int(df.iloc[0, 1])
        for i in range(num_scenarios):
            if is_number(df.iloc[0, i + 2]):
                prob_scenarios.append(float(df.iloc[0, i + 2]))
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(1)

    if num_scenarios != len(prob_scenarios):
        print('[WARNING] EnergyStorage file. Number of scenarios different from the probability vector!')

    if round(sum(prob_scenarios), 2) != 1.00:
        print('[ERROR] Probability of scenarios does not add up to 100%. Check file {}. Exiting.'.format(filename))
        exit(ERROR_MARKET_DATA_FILE)

    return num_scenarios, prob_scenarios


def _get_market_costs_from_excel_file(filename, sheet_name, num_scenarios):
    data = pd.read_excel(filename, sheet_name=sheet_name)
    _, num_cols = data.shape
    cost_values = dict()
    scn_idx = 0
    for i in range(num_scenarios):
        cost_values_scenario = list()
        for j in range(num_cols - 1):
            cost_values_scenario.append(float(data.iloc[i, j + 1]))
        cost_values[scn_idx] = cost_values_scenario
        scn_idx = scn_idx + 1
    return cost_values


# ======================================================================================================================
#  RESULTS PROCESSING functions
# ======================================================================================================================
def _process_operational_planning_results(planning_problem, tso_model, dso_models, esso_model, optimization_results):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()
    processed_results['esso'] = dict()
    processed_results['interface'] = dict()

    processed_results['tso'] = transmission_network.process_results(tso_model, optimization_results['tso'])
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results(dso_model, optimization_results['dso'][node_id])
    processed_results['esso'] = shared_ess_data.process_results(esso_model, optimization_results['esso'])
    processed_results['interface'] = _process_results_interface_power_flow(planning_problem, tso_model, dso_models)

    return processed_results


def _process_results_interface_power_flow(planning_problem, tso_model, dso_models):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results_interface_power_flow(tso_model)
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results_interface_power_flow(dso_model)

    return processed_results


def _process_operational_planning_results_no_coordination(planning_problem, tso_model, dso_models, optimization_results):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results(tso_model, optimization_results['tso'])
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results(dso_model, optimization_results['dso'][node_id])

    return processed_results


# ======================================================================================================================
#  RESULTS PLANNING - write functions
# ======================================================================================================================
def _write_planning_results_to_excel(planning_problem, shared_ess_processed_results, shared_ess_capacity, operational_planning_processed_results=dict(), bound_evolution=dict(), filename='planing_results'):

    wb = Workbook()

    # Planning results
    #_write_main_info_to_excel(planning_problem.shared_ess_data, wb, shared_ess_processed_results)
    _write_ess_capacity_investment_to_excel(planning_problem.shared_ess_data, wb, shared_ess_capacity['investment'])
    _write_ess_capacity_available_to_excel(planning_problem.shared_ess_data, wb, shared_ess_capacity['available'])
    _write_secondary_reserve_bands_to_excel(planning_problem.shared_ess_data, wb, shared_ess_processed_results['results'])

    if bound_evolution:
        _write_bound_evolution_to_excel(wb, bound_evolution)

    # Operational Planning Results
    if operational_planning_processed_results:
        _write_operational_planning_main_info_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_shared_energy_storages_results_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_interface_power_flow_results_to_excel(planning_problem, wb, operational_planning_processed_results['interface'])
        _write_network_voltage_results_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_network_consumption_results_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_network_generation_results_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_network_branch_results_to_excel(planning_problem, wb, operational_planning_processed_results, 'losses')
        _write_network_branch_results_to_excel(planning_problem, wb, operational_planning_processed_results, 'ratio')
        _write_network_branch_results_to_excel(planning_problem, wb, operational_planning_processed_results, 'current_perc')
        _write_network_branch_power_flow_results_to_excel(planning_problem, wb, operational_planning_processed_results)
        _write_network_energy_storages_results_to_excel(planning_problem, wb, operational_planning_processed_results)

    results_filename = os.path.join(planning_problem.results_dir, filename + '.xlsx')
    try:
        wb.save(results_filename)
        print(f'[INFO] ESS Optimization Results written to {results_filename}.')
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(planning_problem.results_dir, f'{filename}_{current_time}.xlsx')
        print(f'[INFO] ESS Optimization Results written to {backup_filename}.')
        wb.save(backup_filename)


def _write_main_info_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'

    # Objective function value
    line = 1
    sheet.cell(row=line, column=1).value = 'Objective (cost) [Mm.u.]'
    sheet.cell(row=line, column=2).value = results['of_value'] / 1e6
    sheet.cell(row=line, column=2).number_format = decimal_style

    # Execution time
    line += 1
    sheet.cell(row=line, column=1).value = 'Execution time, [s]'
    sheet.cell(row=line, column=2).value = results['runtime']
    sheet.cell(row=line, column=2).number_format = decimal_style

    # Number of years
    line += 1
    sheet.cell(row=line, column=1).value = 'Number of years'
    sheet.cell(row=line, column=2).value = len(shared_ess_data.years)

    # Number of representative days
    line += 1
    sheet.cell(row=line, column=1).value = 'Number of days'
    sheet.cell(row=line, column=2).value = len(shared_ess_data.days)

    # Number of price (market) scenarios
    line += 1
    sheet.cell(row=line, column=1).value = 'Number of market scenarios'
    sheet.cell(row=line, column=2).value = len(shared_ess_data.prob_market_scenarios)

    # Number of operation (reserve activation) scenarios
    line += 1
    sheet.cell(row=line, column=1).value = 'Number of operation scenarios (Shared ESS)'
    sheet.cell(row=line, column=2).value = len(shared_ess_data.prob_operation_scenarios)


def _write_ess_capacity_investment_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Capacity Investment')

    years = [year for year in shared_ess_data.years]

    num_style = '0.00'

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Node'
    sheet.cell(row=line_idx, column=2).value = 'Quantity'
    for y in range(len(years)):
        year = years[y]
        sheet.cell(row=line_idx, column=y + 3).value = int(year)

    # Write investment values, power and energy
    for node_id in results:

        # Power capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'S, [MVA]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Energy capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'E, [MVAh]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Power capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost S, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_s = shared_ess_data.cost_investment['power_capacity'][year] * results[node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 3).value = cost_s
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Energy capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost E, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_e = shared_ess_data.cost_investment['energy_capacity'][year] * results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).value = cost_e
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Total capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost Total, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_s = shared_ess_data.cost_investment['power_capacity'][year] * results[node_id][year]['power']
            cost_e = shared_ess_data.cost_investment['energy_capacity'][year] * results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).value = cost_s + cost_e
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style


def _write_ess_capacity_available_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Capacity Available')

    num_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    row_idx, col_idx = 1, 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Node'
    col_idx = col_idx + 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Quantity'
    col_idx = col_idx + 1
    for year in shared_ess_data.years:
        sheet.cell(row=row_idx, column=col_idx).value = int(year)
        col_idx = col_idx + 1

    # Write investment values, power and energy
    for node_id in results:

        # Power capacity
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'S, [MVA]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['power']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Energy capacity
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'E, [MVAh]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['energy']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Degradation factor
        if "degradation_factor" in results[node_id][year]:
            col_idx = 1
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=col_idx).value = node_id
            col_idx = col_idx + 1
            sheet.cell(row=row_idx, column=col_idx).value = 'Degradation factor'
            col_idx = col_idx + 1
            for year in shared_ess_data.years:
                sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['degradation_factor']
                sheet.cell(row=row_idx, column=col_idx).number_format = perc_style
                col_idx = col_idx + 1


def _write_secondary_reserve_bands_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('ESS, Secondary Reserve')

    num_style = '0.00'
    repr_years = [year for year in shared_ess_data.years]
    repr_days = [day for day in shared_ess_data.days]

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Year'
    sheet.cell(row=line_idx, column=2).value = 'Day'
    sheet.cell(row=line_idx, column=3).value = 'Type'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=line_idx, column=p + 4).value = p

    for y in range(len(shared_ess_data.years)):
        year = repr_years[y]
        for d in range(len(repr_days)):
            day = repr_days[d]
            pup_total = [0.0] * shared_ess_data.num_instants
            pdown_total = [0.0] * shared_ess_data.num_instants
            for e in range(len(shared_ess_data.active_distribution_network_nodes)):
                node_id = shared_ess_data.active_distribution_network_nodes[e]
                for s_m in range(len(shared_ess_data.prob_market_scenarios)):
                    prob_market_scn = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in range(len(shared_ess_data.prob_operation_scenarios)):
                        prob_oper_scn = shared_ess_data.prob_operation_scenarios[s_o]
                        for p in range(shared_ess_data.num_instants):
                            pup = results[year][day][s_m][s_o]['p_up'][node_id][p]
                            pdown = results[year][day][s_m][s_o]['p_down'][node_id][p]
                            if pup != 'N/A':
                                pup_total[p] += pup * prob_market_scn * prob_oper_scn
                            if pdown != 'N/A':
                                pdown_total[p] += pdown * prob_market_scn * prob_oper_scn

            # Upward reserve - per day
            line_idx += 1
            sheet.cell(row=line_idx, column=1).value = int(year)
            sheet.cell(row=line_idx, column=2).value = day
            sheet.cell(row=line_idx, column=3).value = 'Upward, [MW]'
            for p in range(shared_ess_data.num_instants):
                sheet.cell(row=line_idx, column=p + 4).value = pup_total[p]
                sheet.cell(row=line_idx, column=p + 4).number_format = num_style

            # Downward reserve - per day
            line_idx += 1
            sheet.cell(row=line_idx, column=1).value = int(year)
            sheet.cell(row=line_idx, column=2).value = day
            sheet.cell(row=line_idx, column=3).value = 'Downward, [MW]'
            for p in range(shared_ess_data.num_instants):
                sheet.cell(row=line_idx, column=p + 4).value = -pdown_total[p]
                sheet.cell(row=line_idx, column=p + 4).number_format = num_style


def _write_bound_evolution_to_excel(workbook, bound_evolution):

    sheet = workbook.create_sheet('Convergence Characteristic')

    lower_bound = bound_evolution['lower_bound']
    upper_bound = bound_evolution['upper_bound']
    num_lines = max(len(upper_bound), len(lower_bound))

    num_style = '0.00'

    # Write header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Iteration'
    sheet.cell(row=line_idx, column=2).value = 'Lower Bound, [NPV Mm.u.]'
    sheet.cell(row=line_idx, column=3).value = 'Upper Bound, [NPV Mm.u.]'

    # Iterations
    line_idx = 2
    for i in range(num_lines):
        sheet.cell(row=line_idx, column=1).value = i
        line_idx += 1

    # Lower bound
    line_idx = 2
    for value in lower_bound:
        sheet.cell(row=line_idx, column=2).value = value / 1e6
        sheet.cell(row=line_idx, column=2).number_format = num_style
        line_idx += 1

    # Upper bound
    line_idx = 2
    for value in upper_bound:
        sheet.cell(row=line_idx, column=3).value = value / 1e6
        sheet.cell(row=line_idx, column=3).number_format = num_style
        line_idx += 1


# ======================================================================================================================
#  RESULTS OPERATIONAL PLANNING - write functions
# ======================================================================================================================
def _write_operational_planning_results_to_excel(planning_problem, results, primal_evolution=list(), filename='operation_planning_results'):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(wb, results)
    _write_shared_ess_specifications(wb, planning_problem.shared_ess_data)

    if primal_evolution:
        _write_objective_function_evolution_to_excel(wb, primal_evolution)

    # Interface Power Flow
    _write_interface_power_flow_results_to_excel(planning_problem, wb, results['interface'])

    # Shared Energy Storages results
    _write_shared_energy_storages_results_to_excel(planning_problem, wb, results)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'current_perc')

    # Save results
    results_filename = os.path.join(planning_problem.results_dir, filename + '.xlsx')
    try:
        wb.save(results_filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(planning_problem.results_dir, f'{filename}_{current_time}.xlsx')
        print(f"[WARNING] Couldn't write to file {results_filename}. Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_results_no_coordination_to_excel(planning_problem, results, filename='operation_planning_results_no_coordination'):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'current_perc')
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)

    # Save results
    results_filename = os.path.join(planning_problem.results_dir, filename + '.xlsx')
    try:
        wb.save(results_filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(planning_problem.results_dir, f'{filename}_{current_time}.xlsx')
        print(f"[WARNING] Couldn't write to file {results_filename}. Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_main_info_to_excel(planning_problem, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'
    line_idx = 1

    # Write Header
    col_idx = 4
    for year in planning_problem.years:
        for _ in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1

    col_idx = 1
    line_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Agent'
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Node ID'
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Value'
    col_idx += 1

    for _ in planning_problem.years:
        for day in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1

    # ESSO
    if 'esso' in results:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = 'ESSO'
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = '-'
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Objective (cost), [€]'
        col_idx += 1
        for year in results['esso']['results']:
            for day in results['esso']['results'][year]:
                sheet.cell(row=line_idx, column=col_idx).value = results['esso']['results'][year][day]['obj']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # TSO
    line_idx = _write_operational_planning_main_info_per_operator(planning_problem.transmission_network, sheet, 'TSO', line_idx, results['tso']['results'])

    # DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id]
        line_idx = _write_operational_planning_main_info_per_operator(distribution_network, sheet, 'DSO', line_idx, dso_results, tn_node_id=tn_node_id)


def _write_operational_planning_main_info_per_operator(network, sheet, operator_type, line_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1

    # - Objective
    obj_string = 'Objective'
    if network.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=col_idx).value = obj_string
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Load
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Load, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            load_aux = results[year][day]['total_load']
            if network.params.l_curt:
                load_aux -= results[year][day]['load_curt']
            sheet.cell(row=line_idx, column=col_idx).value = load_aux
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Flexibility used
    if network.params.fl_reg:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Flexibility used, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['flex_used']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Total Load curtailed
    if network.params.l_curt:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Load curtailed, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['load_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Total Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Conventional Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Conventional Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_conventional_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Renewable Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_renewable_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Renewable Generation Curtailed
    if network.params.rg_curt:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation curtailed, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['gen_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Losses
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Losses, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['losses']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Number of price (market) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of market scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_market_scenarios)
            col_idx += 1

    # Number of operation (generation and consumption) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of operation scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_operation_scenarios)
            col_idx += 1

    return line_idx


def _write_shared_ess_specifications(workbook, shared_ess_info):

    sheet = workbook.create_sheet('Shared ESS Specifications')

    decimal_style = '0.000'

    # Write Header
    row_idx = 1
    sheet.cell(row=row_idx, column=1).value = 'Year'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Sinst, [MVA]'
    sheet.cell(row=row_idx, column=4).value = 'Einst, [MVAh]'

    # Write Shared ESS specifications
    for year in shared_ess_info.years:
        for shared_ess in shared_ess_info.shared_energy_storages[year]:
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = year
            sheet.cell(row=row_idx, column=2).value = shared_ess.bus
            sheet.cell(row=row_idx, column=3).value = shared_ess.s
            sheet.cell(row=row_idx, column=3).number_format = decimal_style
            sheet.cell(row=row_idx, column=4).value = shared_ess.e
            sheet.cell(row=row_idx, column=4).number_format = decimal_style


def _write_objective_function_evolution_to_excel(workbook, primal_evolution):

    sheet = workbook.create_sheet('Primal Evolution')

    decimal_style = '0.000000'
    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Iteration'
    sheet.cell(row=row_idx, column=2).value = 'OF value'
    row_idx = row_idx + 1
    for i in range(len(primal_evolution)):
        sheet.cell(row=row_idx, column=1).value = i
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        row_idx = row_idx + 1


def _write_interface_power_flow_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Interface PF')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    # TSO's results
    for year in results['tso']:
        for day in results['tso'][year]:
            for node_id in results['tso'][year][day]:
                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q = [0.0 for _ in range(planning_problem.num_instants)]
                for s_m in results['tso'][year][day][node_id]:
                    omega_m = planning_problem.transmission_network.network[year][day].prob_market_scenarios[s_m]
                    for s_o in results['tso'][year][day][node_id][s_m]:
                        omega_s = planning_problem.transmission_network.network[year][day].prob_operation_scenarios[s_o]

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_p = results['tso'][year][day][node_id][s_m][s_o]['p'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_p[p] += interface_p * omega_m * omega_s
                        row_idx += 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_q = results['tso'][year][day][node_id][s_m][s_o]['q'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_q[p] += interface_q * omega_m * omega_s
                        row_idx += 1

                # Expected Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

    # DSOs' results
    for node_id in results['dso']:
        for year in results['dso'][node_id]:
            for day in results['dso'][node_id][year]:
                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q = [0.0 for _ in range(planning_problem.num_instants)]
                for s_m in results['dso'][node_id][year][day]:
                    omega_m = planning_problem.distribution_networks[node_id].network[year][day].prob_market_scenarios[s_m]
                    for s_o in results['dso'][node_id][year][day][s_m]:
                        omega_s = planning_problem.distribution_networks[node_id].network[year][day].prob_operation_scenarios[s_o]

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_p = results['dso'][node_id][year][day][s_m][s_o]['p'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_p[p] += interface_p * omega_m * omega_s
                        row_idx += 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(len(results['dso'][node_id][year][day][s_m][s_o]['q'])):
                            interface_q = results['dso'][node_id][year][day][s_m][s_o]['q'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_q[p] += interface_q * omega_m * omega_s
                        row_idx += 1

                # Expected Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(len(results['dso'][node_id][year][day][s_m][s_o]['q'])):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1


def _write_shared_energy_storages_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Shared ESS')

    row_idx = 1
    decimal_style = '0.00'
    percent_style = '0.00%'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p

    # ESSO's results
    for year in results['esso']['results']:
        for day in results['esso']['results'][year]:

            expected_p = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            expected_pup = dict()
            expected_pdown = dict()
            for node_id in planning_problem.active_distribution_network_nodes:
                expected_p[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_pup[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_pdown[node_id] = [0.0 for _ in range(planning_problem.num_instants)]

            for s_m in results['esso']['results'][year][day]:
                if s_m != 'obj':

                    omega_m = planning_problem.shared_ess_data.prob_market_scenarios[s_m]

                    for s_o in results['esso']['results'][year][day][s_m]:

                        omega_s = planning_problem.shared_ess_data.prob_operation_scenarios[s_o]

                        for node_id in planning_problem.active_distribution_network_nodes:

                            # Active power
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'ESSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_p = results['esso']['results'][year][day][s_m][s_o]['p'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_p
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_p != 'N/A':
                                    expected_p[node_id][p] += ess_p * omega_m * omega_s
                                else:
                                    expected_p[node_id][p] = ess_p

                            # State-of-Charge, [MVAh]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'ESSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc = results['esso']['results'][year][day][s_m][s_o]['soc'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_soc != 'N/A':
                                    expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                                else:
                                    expected_soc[node_id][p] = ess_soc

                            # State-of-Charge, [%]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'ESSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc_percent = results['esso']['results'][year][day][s_m][s_o]['soc_percent'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc_percent
                                sheet.cell(row=row_idx, column=p + 8).number_format = percent_style
                                if ess_soc_percent != 'N/A':
                                    expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                                else:
                                    expected_soc_percent[node_id][p] = ess_soc_percent

                            # Secondary reserve - Upward band, [MW]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'ESSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Pup, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_pup = results['esso']['results'][year][day][s_m][s_o]['p_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_pup
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_pup != 'N/A':
                                    expected_pup[node_id][p] += ess_pup * omega_m * omega_s
                                else:
                                    expected_pup[node_id][p] = ess_pup

                            # Secondary reserve - Downward band, [MW]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'ESSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Pdown, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_pdown = results['esso']['results'][year][day][s_m][s_o]['p_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_pdown
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_pup != 'N/A':
                                    expected_pdown[node_id][p] += ess_pdown * omega_m * omega_s
                                else:
                                    expected_pdown[node_id][p] = ess_pdown

            for node_id in planning_problem.active_distribution_network_nodes:

                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = percent_style

                # Secondary reserve - Upward band, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Pup, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_pup[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Secondary reserve - Downward band, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Pdown, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_pdown[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

    # TSO's results
    for year in results['tso']['results']:
        for day in results['tso']['results'][year]:

            expected_p = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for node_id in planning_problem.active_distribution_network_nodes:
                expected_p[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent[node_id] = [0.0 for _ in range(planning_problem.num_instants)]

            for s_m in results['tso']['results'][year][day]:
                if s_m not in exclusions:

                    omega_m = planning_problem.transmission_network.network[year][day].prob_market_scenarios[s_m]

                    for s_o in results['tso']['results'][year][day][s_m]:

                        omega_s = planning_problem.transmission_network.network[year][day].prob_operation_scenarios[s_o]

                        for node_id in planning_problem.active_distribution_network_nodes:

                            # Active power
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'TSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_p = results['tso']['results'][year][day][s_m][s_o]['shared_energy_storages']['p'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_p
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_p != 'N/A':
                                    expected_p[node_id][p] += ess_p * omega_m * omega_s
                                else:
                                    expected_p[node_id][p] = ess_p

                            # State-of-Charge, [MVAh]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'TSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc = results['tso']['results'][year][day][s_m][s_o]['shared_energy_storages']['soc'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_soc != 'N/A':
                                    expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                                else:
                                    expected_soc[node_id][p] = ess_soc

                            # State-of-Charge, [%]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'TSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc_percent = results['tso']['results'][year][day][s_m][s_o]['shared_energy_storages']['soc_percent'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc_percent
                                sheet.cell(row=row_idx, column=p + 8).number_format = percent_style
                                if ess_soc_percent != 'N/A':
                                    expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                                else:
                                    expected_soc_percent[node_id][p] = ess_soc_percent

            for node_id in planning_problem.active_distribution_network_nodes:

                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = percent_style

    # DSO's results
    for node_id in results['dso']:
        for year in results['dso'][node_id]['results']:
            for day in results['dso'][node_id]['results'][year]:

                distribution_network = planning_problem.distribution_networks[node_id].network[year][day]
                ref_node_id = distribution_network.get_reference_node_id()

                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent = [0.0 for _ in range(planning_problem.num_instants)]

                for s_m in results['dso'][node_id]['results'][year][day]:
                    if s_m not in exclusions:

                        omega_m = distribution_network.prob_market_scenarios[s_m]

                        for s_o in results['dso'][node_id]['results'][year][day][s_m]:

                            omega_s = distribution_network.prob_operation_scenarios[s_o]

                            # Active power
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'DSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_p = results['dso'][node_id]['results'][year][day][s_m][s_o]['shared_energy_storages']['p'][ref_node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_p
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_p != 'N/A':
                                    expected_p[p] += ess_p * omega_m * omega_s
                                else:
                                    expected_p[p] = ess_p

                            # State-of-Charge, [MVAh]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'DSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc = results['dso'][node_id]['results'][year][day][s_m][s_o]['shared_energy_storages']['soc'][ref_node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if ess_soc != 'N/A':
                                    expected_soc[p] += ess_soc * omega_m * omega_s
                                else:
                                    expected_soc[p] = ess_soc

                            # State-of-Charge, [%]
                            row_idx = row_idx + 1
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = 'DSO'
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(planning_problem.num_instants):
                                ess_soc_percent = results['dso'][node_id]['results'][year][day][s_m][s_o]['shared_energy_storages']['soc_percent'][ref_node_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = ess_soc_percent
                                sheet.cell(row=row_idx, column=p + 8).number_format = percent_style
                                if ess_soc_percent != 'N/A':
                                    expected_soc_percent[p] += ess_soc_percent * omega_m * omega_s
                                else:
                                    expected_soc_percent[p] = ess_soc_percent

                # Expected values
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_percent[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = percent_style


def _write_network_voltage_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_voltage_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_voltage_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_voltage_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_vmag = dict()
            expected_vang = dict()
            for node in network[year][day].nodes:
                expected_vmag[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_vang[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for node_id in results[year][day][s_m][s_o]['voltage']['vmag']:

                            v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                            # Voltage magnitude
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Vmag, [p.u.]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                v_mag = results[year][day][s_m][s_o]['voltage']['vmag'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = v_mag
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                if v_mag > v_max + SMALL_TOLERANCE or v_mag < v_min - SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                                expected_vmag[node_id][p] += v_mag * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Voltage angle
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Vang, [º]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                v_ang = results[year][day][s_m][s_o]['voltage']['vang'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = v_ang
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_vang[node_id][p] += v_mag * omega_m * omega_s
                            row_idx = row_idx + 1

            for node in network[year][day].nodes:

                node_id = node.bus_i
                v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                # Expected voltage magnitude
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_vmag[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    if expected_vmag[node_id][p] > v_max + SMALL_TOLERANCE or expected_vmag[node_id][p] < v_min - SMALL_TOLERANCE:
                        sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                row_idx = row_idx + 1

                # Expected voltage angle
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Vang, [º]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_vang[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_consumption_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']['results']
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    row_idx = _write_network_consumption_results_per_operator(transmission_network, tn_params, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        row_idx = _write_network_consumption_results_per_operator(distribution_network, dn_params, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_consumption_results_per_operator(network, params, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_pc = dict()
            expected_flex_up = dict()
            expected_flex_down = dict()
            expected_pc_curt = dict()
            expected_pnet = dict()
            expected_qc = dict()
            for node in network[year][day].nodes:
                expected_pc[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_flex_up[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_flex_down[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pc_curt[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pnet[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qc[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for node_id in results[year][day][s_m][s_o]['consumption']['pc']:

                            # - Active Power
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pc, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                pc = results[year][day][s_m][s_o]['consumption']['pc'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pc
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_pc[node_id][p] += pc * omega_m * omega_s
                            row_idx = row_idx + 1

                            if params.fl_reg:

                                # - Flexibility, up
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Flex Up, [MW]'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    flex = results[year][day][s_m][s_o]['consumption']['p_up'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).value = flex
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                    expected_flex_up[node_id][p] += flex * omega_m * omega_s
                                row_idx = row_idx + 1

                                # - Flexibility, down
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Flex Down, [MW]'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    flex = results[year][day][s_m][s_o]['consumption']['p_down'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).value = flex
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                    expected_flex_down[node_id][p] += flex * omega_m * omega_s
                                row_idx = row_idx + 1

                            if params.l_curt:

                                # - Active power curtailment
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Pc_curt, [MW]'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    pc_curt = results[year][day][s_m][s_o]['consumption']['pc_curt'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).value = pc_curt
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                    if pc_curt >= SMALL_TOLERANCE:
                                        sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                                    expected_pc_curt[node_id][p] += pc_curt * omega_m * omega_s
                                row_idx = row_idx + 1

                            if params.fl_reg or params.l_curt:

                                # - Active power net consumption
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Pc_net, [MW]'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    p_net = results[year][day][s_m][s_o]['consumption']['pc_net'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).value = p_net
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                    expected_pnet[node_id][p] += p_net * omega_m * omega_s
                                row_idx = row_idx + 1

                            # - Reactive power
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Qc, [MVAr]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                qc = results[year][day][s_m][s_o]['consumption']['qc'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = qc
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_qc[node_id][p] += qc * omega_m * omega_s
                            row_idx = row_idx + 1

            for node in network[year][day].nodes:

                node_id = node.bus_i

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Pc, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_pc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                if params.fl_reg:

                    # - Flexibility, up
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Flex Up, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_flex_up[node_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                    # - Flexibility, down
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Flex Down, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_flex_down[node_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                if params.l_curt:

                    # - Load curtailment (active power)
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pc_curt, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pc_curt[node_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                        if expected_pc_curt[node_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                    row_idx = row_idx + 1

                if params.fl_reg or params.l_curt:

                    # - Active power net consumption
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pc_net, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pnet[node_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Reactive power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Qc, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_qc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_generation_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Generator ID'
    sheet.cell(row=row_idx, column=5).value = 'Type'
    sheet.cell(row=row_idx, column=6).value = 'Year'
    sheet.cell(row=row_idx, column=7).value = 'Day'
    sheet.cell(row=row_idx, column=8).value = 'Quantity'
    sheet.cell(row=row_idx, column=9).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=10).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 11).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    row_idx = _write_network_generation_results_per_operator(transmission_network, tn_params, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        row_idx = _write_network_generation_results_per_operator(distribution_network, dn_params, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_generation_results_per_operator(network, params, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_pg = dict()
            expected_pg_curt = dict()
            expected_pg_net = dict()
            expected_qg = dict()
            for generator in network[year][day].generators:
                expected_pg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pg_curt[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pg_net[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for g in results[year][day][s_m][s_o]['generation']['pg']:

                            node_id = network[year][day].generators[g].bus
                            gen_id = network[year][day].generators[g].gen_id
                            gen_type = network[year][day].get_gen_type(gen_id)

                            # Active Power
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = gen_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Pg, [MW]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                pg = results[year][day][s_m][s_o]['generation']['pg'][g][p]
                                sheet.cell(row=row_idx, column=p + 11).value = pg
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_pg[gen_id][p] += pg * omega_m * omega_s
                            row_idx = row_idx + 1

                            if params.rg_curt:

                                # Active Power curtailment
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = gen_id
                                sheet.cell(row=row_idx, column=5).value = gen_type
                                sheet.cell(row=row_idx, column=6).value = int(year)
                                sheet.cell(row=row_idx, column=7).value = day
                                sheet.cell(row=row_idx, column=8).value = 'Pg_curt, [MW]'
                                sheet.cell(row=row_idx, column=9).value = s_m
                                sheet.cell(row=row_idx, column=10).value = s_o
                                for p in range(network[year][day].num_instants):
                                    pg_curt = results[year][day][s_m][s_o]['generation']['pg_curt'][g][p]
                                    sheet.cell(row=row_idx, column=p + 11).value = pg_curt
                                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                    if pg_curt > SMALL_TOLERANCE:
                                        sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                                    expected_pg_curt[gen_id][p] += pg_curt * omega_m * omega_s
                                row_idx = row_idx + 1

                                # Active Power net
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = gen_id
                                sheet.cell(row=row_idx, column=5).value = gen_type
                                sheet.cell(row=row_idx, column=6).value = int(year)
                                sheet.cell(row=row_idx, column=7).value = day
                                sheet.cell(row=row_idx, column=8).value = 'Pg_net, [MW]'
                                sheet.cell(row=row_idx, column=9).value = s_m
                                sheet.cell(row=row_idx, column=10).value = s_o
                                for p in range(network[year][day].num_instants):
                                    pg_net = results[year][day][s_m][s_o]['generation']['pg_net'][g][p]
                                    sheet.cell(row=row_idx, column=p + 11).value = pg_net
                                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                    expected_pg_net[gen_id][p] += pg_net * omega_m * omega_s
                                row_idx = row_idx + 1

                            # Reactive Power
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = gen_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Qg, [MVAr]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                qg = results[year][day][s_m][s_o]['generation']['qg'][g][p]
                                sheet.cell(row=row_idx, column=p + 11).value = qg
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_qg[gen_id][p] += qg * omega_m * omega_s
                            row_idx = row_idx + 1

            for generator in network[year][day].generators:

                node_id = generator.bus
                gen_id = generator.gen_id
                gen_type = network[year][day].get_gen_type(gen_id)

                # Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = gen_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Pg, [MW]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_pg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                if params.rg_curt:

                    # Active Power curtailment
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = gen_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Pg_curt, [MW]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_pg_curt[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                        if expected_pg_curt[gen_id][p] > SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                    row_idx = row_idx + 1

                    # Active Power net
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = node_id
                    sheet.cell(row=row_idx, column=4).value = gen_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Pg_net, [MW]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_pg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

                # Reactive Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = gen_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Qg, [MVAr]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_qg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_branch_results_to_excel(planning_problem, workbook, results, result_type):

    sheet_name = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'
    elif result_type == 'current_perc':
        sheet_name = 'Branch Loading'
    sheet = workbook.create_sheet(sheet_name)

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'From Node ID'
    sheet.cell(row=row_idx, column=4).value = 'To Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Year'
    sheet.cell(row=row_idx, column=6).value = 'Day'
    sheet.cell(row=row_idx, column=7).value = 'Quantity'
    sheet.cell(row=row_idx, column=8).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=9).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 10).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_branch_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'], result_type)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_branch_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, result_type, tn_node_id=tn_node_id)


def _write_network_branch_results_per_operator(network, sheet, operator_type, row_idx, results, result_type, tn_node_id='-'):

    decimal_style = '0.00'
    perc_style = '0.00%'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    aux_string = str()
    if result_type == 'losses':
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        aux_string = 'Ratio'
    elif result_type == 'current_perc':
        aux_string = 'I, [%]'

    for year in results:
        for day in results[year]:

            expected_values = dict()
            for k in range(len(network[year][day].branches)):
                expected_values[k] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for k in results[year][day][s_m][s_o]['branches'][result_type]:
                            branch = network[year][day].branches[k]
                            if not(result_type == 'ratio' and not branch.is_transformer):

                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = branch.fbus
                                sheet.cell(row=row_idx, column=4).value = branch.tbus
                                sheet.cell(row=row_idx, column=5).value = int(year)
                                sheet.cell(row=row_idx, column=6).value = day
                                sheet.cell(row=row_idx, column=7).value = aux_string
                                sheet.cell(row=row_idx, column=8).value = s_m
                                sheet.cell(row=row_idx, column=9).value = s_o
                                for p in range(network[year][day].num_instants):
                                    value = results[year][day][s_m][s_o]['branches'][result_type][k][p]
                                    if result_type == 'current_perc':
                                        sheet.cell(row=row_idx, column=p + 10).value = value
                                        sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                                        if value > 1.0 + SMALL_TOLERANCE:
                                            sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                                    else:
                                        sheet.cell(row=row_idx, column=p + 10).value = value
                                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                    expected_values[k][p] += value * omega_m * omega_s
                                row_idx = row_idx + 1

            for k in range(len(network[year][day].branches)):
                branch = network[year][day].branches[k]
                if not (result_type == 'ratio' and not branch.is_transformer):

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = branch.fbus
                    sheet.cell(row=row_idx, column=4).value = branch.tbus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = aux_string
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        if result_type == 'current_perc':
                            sheet.cell(row=row_idx, column=p + 10).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            if expected_values[k][p] > 1.0:
                                sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                        else:
                            sheet.cell(row=row_idx, column=p + 10).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                    row_idx = row_idx + 1

    return row_idx


def _write_network_branch_power_flow_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'From Node ID'
    sheet.cell(row=row_idx, column=4).value = 'To Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Year'
    sheet.cell(row=row_idx, column=6).value = 'Day'
    sheet.cell(row=row_idx, column=7).value = 'Quantity'
    sheet.cell(row=row_idx, column=8).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=9).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 10).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_power_flow_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_power_flow_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_power_flow_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    perc_style = '0.00%'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']

    for year in results:
        for day in results[year]:

            expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
            for k in range(len(network[year][day].branches)):
                expected_values['pij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['pji'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qji'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sji'][k] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for k in range(len(network[year][day].branches)):

                            branch = network[year][day].branches[k]
                            rating = branch.rate_a
                            if rating == 0.0:
                                rating = BRANCH_UNKNOWN_RATING

                            # Pij, [MW]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['pij'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['pij'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Pij, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'P, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['pij'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

                            # Pji, [MW]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['pji'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['pji'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Pji, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'P, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['pji'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

                            # Qij, [MVAr]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['qij'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['qij'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Qij, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Q, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['qij'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

                            # Qji, [MW]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['qji'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['qji'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Qji, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Q, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['qji'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

                            # Sij, [MVA]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['sij'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['sij'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Sij, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'S, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['sij'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

                            # Sji, [MW]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day][s_m][s_o]['branches']['power_flow']['sji'][k][p]
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_values['sji'][k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Sji, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.fbus
                            sheet.cell(row=row_idx, column=4).value = branch.tbus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'S, [%]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = abs(results[year][day][s_m][s_o]['branches']['power_flow']['sji'][k][p] / rating)
                                sheet.cell(row=row_idx, column=p + 10).value = value
                                sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                            row_idx = row_idx + 1

            for k in range(len(network[year][day].branches)):

                branch = network[year][day].branches[k]
                rating = branch.rate_a
                if rating == 0.0:
                    rating = BRANCH_UNKNOWN_RATING

                # Pij, [MW]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['pij'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Pij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'P, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['pij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                row_idx = row_idx + 1

                # Pji, [MW]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['pji'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Pji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'P, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['pji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                row_idx = row_idx + 1

                # Qij, [MVAr]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['qij'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Qij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['qij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                row_idx = row_idx + 1

                # Qji, [MVAr]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['qji'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Qji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['qji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [MVA]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['sij'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'S, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['sij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                row_idx = row_idx + 1

                # Sji, [MVA]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_values['sji'][k][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # Sji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'S, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = abs(expected_values['sji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 10).number_format = perc_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_energy_storages_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'Connection Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']['results']
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    row_idx = _write_network_energy_storages_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        row_idx = _write_network_energy_storages_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_energy_storages_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    percent_style = '0.00%'
    exclusions = ['runtime', 'obj', 'gen_cost', 'total_load', 'total_gen', 'total_conventional_gen', 'total_renewable_gen', 'losses', 'gen_curt', 'load_curt', 'flex_used']

    for year in results:
        for day in results[year]:

            expected_p = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for energy_storage in network[year][day].energy_storages:
                expected_p[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc_percent[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]:
                if s_m not in exclusions:
                    omega_m = network[year][day].prob_market_scenarios[s_m]
                    for s_o in results[year][day][s_m]:
                        omega_s = network[year][day].prob_operation_scenarios[s_o]
                        for node_id in results[year][day][s_m][s_o]['energy_storages']['p']:

                            # - Active Power
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                ess_p = results[year][day][s_m][s_o]['energy_storages']['p'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = ess_p
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_p[node_id][p] += ess_p * omega_m * omega_s
                            row_idx = row_idx + 1

                            # State-of-Charge, [MWh]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'SoC, [MWh]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                ess_soc = results[year][day][s_m][s_o]['energy_storages']['soc'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = ess_soc
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                if ess_soc != 'N/A':
                                    expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                                else:
                                    expected_soc[node_id][p] = ess_soc
                            row_idx = row_idx + 1

                            # State-of-Charge, [%]
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'SoC, [%]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                ess_soc_percent = results[year][day][s_m][s_o]['energy_storages']['soc_percent'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = ess_soc_percent
                                sheet.cell(row=row_idx, column=p + 9).number_format = percent_style
                                if ess_soc_percent != 'N/A':
                                    expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                                else:
                                    expected_soc_percent[node_id][p] = ess_soc_percent
                            row_idx = row_idx + 1

            for energy_storage in network[year][day].energy_storages:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [MWh]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = percent_style
                row_idx = row_idx + 1

    return row_idx


# ======================================================================================================================
#   NETWORK diagram functions (plot)
# ======================================================================================================================
def _plot_networkx_diagram(planning_problem):

    for year in planning_problem.years:
        for day in planning_problem.days:

            transmission_network = planning_problem.transmission_network.network[year][day]

            node_labels = {}
            ref_nodes, pv_nodes, pq_nodes = [], [], []
            res_pv_nodes = [gen.bus for gen in transmission_network.generators if gen.gen_type == GEN_NONCONVENTIONAL_SOLAR]
            res_wind_nodes = [gen.bus for gen in transmission_network.generators if gen.gen_type == GEN_NONCONVENTIONAL_WIND]
            adn_nodes = planning_problem.active_distribution_network_nodes

            branches = []
            line_list, open_line_list = [], []
            transf_list, open_transf_list = [], []
            for branch in transmission_network.branches:
                if branch.is_transformer:
                    branches.append({'type': 'transformer', 'data': branch})
                else:
                    branches.append({'type': 'line', 'data': branch})

            # Build graph
            graph = nx.Graph()
            for i in range(len(transmission_network.nodes)):
                node = transmission_network.nodes[i]
                graph.add_node(node.bus_i)
                node_labels[node.bus_i] = '{}'.format(node.bus_i)
                if node.type == BUS_REF:
                    ref_nodes.append(node.bus_i)
                elif node.type == BUS_PV:
                    pv_nodes.append(node.bus_i)
                elif node.type == BUS_PQ:
                    if node.bus_i not in (res_pv_nodes + res_wind_nodes + adn_nodes):
                        pq_nodes.append(node.bus_i)
            for i in range(len(branches)):
                branch = branches[i]
                if branch['type'] == 'line':
                    graph.add_edge(branch['data'].fbus, branch['data'].tbus)
                    if branch['data'].status == 1:
                        line_list.append((branch['data'].fbus, branch['data'].tbus))
                    else:
                        open_line_list.append((branch['data'].fbus, branch['data'].tbus))
                if branch['type'] == 'transformer':
                    graph.add_edge(branch['data'].fbus, branch['data'].tbus)
                    if branch['data'].status == 1:
                        transf_list.append((branch['data'].fbus, branch['data'].tbus))
                    else:
                        open_transf_list.append((branch['data'].fbus, branch['data'].tbus))

            # Plot diagram
            pos = nx.spring_layout(graph, k=0.50, iterations=1000)
            fig, ax = plt.subplots(figsize=(12, 8))
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=ref_nodes, node_color='red', node_size=250, label='Reference bus')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=pv_nodes, node_color='lightgreen', node_size=250, label='Conventional generator')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=pq_nodes, node_color='lightblue', node_size=250, label='PQ buses')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=res_pv_nodes, node_color='yellow', node_size=250, label='RES, PV')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=res_wind_nodes, node_color='blue', node_size=250, label='RES, Wind')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=adn_nodes, node_color='orange', node_size=250, label='ADN buses')
            nx.draw_networkx_labels(graph, ax=ax, pos=pos, labels=node_labels, font_size=12)
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=line_list, width=1.50, edge_color='black')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=transf_list, width=2.00, edge_color='blue', label='Transformer')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=open_line_list, style='dashed', width=1.50, edge_color='red')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=open_transf_list, style='dashed', width=2.00, edge_color='red')
            plt.legend(scatterpoints=1, frameon=False, prop={'size': 12})
            plt.axis('off')

            filename = os.path.join(planning_problem.diagrams_dir, f'{planning_problem.name}_{year}_{day}')
            plt.savefig(f'{filename}.pdf', bbox_inches='tight')
            plt.savefig(f'{filename}.png', bbox_inches='tight')


# ======================================================================================================================
#   Aux functions
# ======================================================================================================================
def _get_initial_candidate_solution(planning_problem):
    candidate_solution = {'investment': {}, 'total_capacity': {}}
    for e in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[e]
        candidate_solution['investment'][node_id] = dict()
        candidate_solution['total_capacity'][node_id] = dict()
        for year in planning_problem.years:
            candidate_solution['investment'][node_id][year] = dict()
            candidate_solution['investment'][node_id][year]['s'] = 0.00
            candidate_solution['investment'][node_id][year]['e'] = 0.00
            candidate_solution['total_capacity'][node_id][year] = dict()
            candidate_solution['total_capacity'][node_id][year]['s'] = 0.00
            candidate_solution['total_capacity'][node_id][year]['e'] = 0.00
    return candidate_solution


def _add_shared_energy_storage_to_transmission_network(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            s_base = planning_problem.transmission_network.network[year][day].baseMVA
            for node_id in planning_problem.distribution_networks:
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = node_id
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].name
                shared_energy_storage.s = shared_energy_storage.s / s_base
                shared_energy_storage.e = shared_energy_storage.e / s_base
                planning_problem.transmission_network.network[year][day].shared_energy_storages.append(shared_energy_storage)


def _add_shared_energy_storage_to_distribution_network(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            for node_id in planning_problem.distribution_networks:
                s_base = planning_problem.distribution_networks[node_id].network[year][day].baseMVA
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = planning_problem.distribution_networks[node_id].network[year][day].get_reference_node_id()
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].network[year][day].name
                shared_energy_storage.s = shared_energy_storage.s / s_base
                shared_energy_storage.e = shared_energy_storage.e / s_base
                planning_problem.distribution_networks[node_id].network[year][day].shared_energy_storages.append(shared_energy_storage)
